"""Shared openpyxl styling helpers so every sheet in the generated workbook
looks consistent and "enterprise pricing sheet" professional rather than a
raw data dump."""
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

CURRENCY_FORMAT = '#,##0.00'
PERCENT_FORMAT = '0.00"%"'

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def header_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def light_fill(hex_color: str) -> PatternFill:
    # Lighten a hex color by blending toward white for section/subtotal rows.
    r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
    r = int(r + (255 - r) * 0.85)
    g = int(g + (255 - g) * 0.85)
    b = int(b + (255 - b) * 0.85)
    light_hex = f"{r:02X}{g:02X}{b:02X}"
    return PatternFill(start_color=light_hex, end_color=light_hex, fill_type="solid")


def write_title(ws: Worksheet, text: str, accent_hex: str, row: int = 1, span: int = 6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=16, color=accent_hex)
    ws.row_dimensions[row].height = 24


def write_subtitle(ws: Worksheet, text: str, row: int, span: int = 6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(italic=True, size=10, color="666666")


def write_header_row(ws: Worksheet, row: int, headers: list[str], accent_hex: str):
    fill = header_fill(accent_hex)
    font = Font(bold=True, color="FFFFFF")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = fill
        cell.font = font
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 20


def write_data_row(ws: Worksheet, row: int, values: list, currency_cols: set[int] | None = None, zebra: bool = False):
    currency_cols = currency_cols or set()
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if col_idx in currency_cols and isinstance(value, (int, float)):
            cell.number_format = CURRENCY_FORMAT
        if zebra:
            cell.fill = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")


def autofit_columns(ws: Worksheet, widths: dict[int, int] | None = None, default_min: int = 12, max_width: int = 60):
    computed: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or cell.coordinate in getattr(ws, "merged_cells", []):
                continue
            length = len(str(cell.value))
            col = cell.column
            computed[col] = max(computed.get(col, default_min), min(length + 2, max_width))
    if widths:
        computed.update(widths)
    for col_idx, width in computed.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def freeze_header(ws: Worksheet, header_row: int):
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
