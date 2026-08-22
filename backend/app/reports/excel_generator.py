"""
Excel report generator.

Produces a workbook matching enterprise cloud pricing sheet conventions:
Summary, Compute, Storage, Database, Networking, Licensing, Assumptions,
Validation, Recommendations, Pricing, Totals, Yearly Cost, and Audit sheets -
with color formatting, real Excel SUM formulas (not just pre-computed
numbers), frozen header rows, and a branding header on every sheet.

This module only *renders* an already-computed EstimateResult - it never
recomputes or overrides a price. That guarantees the exported spreadsheet can
never diverge from the numbers the pricing engine produced.
"""
import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from app.domain.branding import BrandingConfig
from app.domain.estimate import EstimateResult
from app.domain.explanation import EstimateExplanation
from app.reports.styles import (
    autofit_columns,
    freeze_header,
    light_fill,
    write_data_row,
    write_header_row,
    write_subtitle,
    write_title,
)


class ExcelReportGenerator:
    def generate(
        self, estimate: EstimateResult, branding: BrandingConfig | None = None,
        explanation: EstimateExplanation | None = None,
    ) -> bytes:
        branding = branding or BrandingConfig()
        wb = Workbook()
        wb.remove(wb.active)

        self._build_summary(wb, estimate, branding, explanation)
        self._build_line_item_sheet(
            wb, "Compute", estimate, branding, {"Compute", "GPU", "Licensing"},
            "Compute Engine, GPU, OS licensing, and GKE cluster/node costs.",
        )
        self._build_line_item_sheet(
            wb, "Storage", estimate, branding, {"Storage"},
            "Persistent Disk and snapshot storage costs.",
        )
        self._build_line_item_sheet(
            wb, "Database", estimate, branding, {"Database"},
            "Cloud SQL compute and storage costs.",
        )
        self._build_line_item_sheet(
            wb, "Networking", estimate, branding, {"Network"},
            "Load balancer and data transfer (egress) costs.",
        )
        self._build_resource_summary(wb, estimate, branding)
        self._build_licensing(wb, estimate, branding)
        self._build_assumptions(wb, estimate, branding, explanation)
        self._build_validation(wb, estimate, branding)
        self._build_recommendations(wb, estimate, branding)
        pricing_last_row = self._build_pricing(wb, estimate, branding)
        self._build_totals(wb, estimate, branding, pricing_last_row)
        self._build_yearly(wb, estimate, branding)
        self._build_audit(wb, estimate, branding)

        wb.active = 0
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # -- Summary ------------------------------------------------------------

    def _build_summary(
        self, wb: Workbook, estimate: EstimateResult, branding: BrandingConfig,
        explanation: EstimateExplanation | None = None,
    ):
        ws = wb.create_sheet("Summary")
        accent = branding.primary_color_hex

        write_title(ws, branding.company_name, accent, row=1)
        write_subtitle(ws, "Google Cloud FinOps Cost Estimate", row=2)

        row = 4
        meta = [
            ("Project", estimate.project_name),
            ("Estimate ID", estimate.estimate_id),
            ("Region", estimate.normalized_spec.region),
            ("Normalization Strategy", estimate.strategy_used.capitalize()),
            ("Prepared By", branding.prepared_by),
            ("Prepared For", branding.prepared_for or "-"),
            ("Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ]
        for label, value in meta:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 1
        write_header_row(ws, row, ["Metric", "Value", "Currency"], accent)
        row += 1
        metrics = [
            ("Monthly Cost", estimate.cost.total_monthly),
            ("Yearly Cost", estimate.cost.total_yearly),
            ("3-Year Cost", estimate.cost.total_three_year),
            ("Monthly Discounts Applied", estimate.cost.discount_total_monthly),
        ]
        for label, value in metrics:
            write_data_row(ws, row, [label, value, estimate.cost.currency], currency_cols={2})
            row += 1

        row += 1
        write_header_row(ws, row, ["Quality Indicator", "Count"], accent)
        row += 1
        indicators = [
            ("Validation Warnings", estimate.validation.warning_count),
            ("Validation Blockers", estimate.validation.blocker_count),
            ("Assumptions Made", len(estimate.assumptions)),
            ("Architecture Components", len(estimate.architecture.components)),
            ("Audit Log Entries", len(estimate.audit_log.entries)),
        ]
        for label, value in indicators:
            write_data_row(ws, row, [label, value])
            row += 1

        if explanation is not None and explanation.executive_summary:
            row += 1
            write_header_row(ws, row, ["AI-Generated Summary"], accent)
            row += 1
            summary_cell = ws.cell(row=row, column=1, value=explanation.executive_summary)
            summary_cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.row_dimensions[row].height = 45
            row += 1

        row += 2
        ws.cell(row=row, column=1, value=branding.footer_note).font = Font(italic=True, size=9, color="888888")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

        autofit_columns(ws, widths={1: 28, 2: 40, 3: 12})
        ws.sheet_view.showGridLines = False

    # -- Generic line-item sheet (Compute / Storage / Database / Networking) --

    def _build_line_item_sheet(
        self, wb: Workbook, title: str, estimate: EstimateResult, branding: BrandingConfig,
        categories: set[str], intro: str,
    ):
        ws = wb.create_sheet(title)
        accent = branding.primary_color_hex
        write_title(ws, title, accent, row=1)
        write_subtitle(ws, intro, row=2)

        items = [li for li in estimate.cost.line_items if li.category in categories]

        header_row = 4
        headers = ["Description", "SKU", "Unit", "Quantity", "Unit Price", "Monthly Amount", "Source"]
        write_header_row(ws, header_row, headers, accent)

        row = header_row + 1
        first_data_row = row
        if not items:
            write_data_row(ws, row, [f"No {title.lower()} resources in this estimate.", "", "", "", "", "", ""])
            row += 1
        else:
            for li in items:
                write_data_row(
                    ws, row,
                    [li.description, li.sku_id or "-", li.unit, li.quantity, li.unit_price, li.monthly_amount, li.source],
                    currency_cols={5, 6},
                )
                row += 1

        if items:
            total_row = row
            ws.cell(row=total_row, column=1, value="Subtotal").font = Font(bold=True)
            total_cell = ws.cell(row=total_row, column=6, value=f"=SUM(F{first_data_row}:F{row - 1})")
            total_cell.font = Font(bold=True)
            total_cell.number_format = "#,##0.00"
            for col in range(1, 8):
                ws.cell(row=total_row, column=col).fill = light_fill(accent)

        freeze_header(ws, header_row)
        autofit_columns(ws, widths={1: 46, 2: 22, 3: 12, 4: 12, 5: 14, 6: 16, 7: 16})
        ws.sheet_view.showGridLines = False

    # -- Resources (per-resource quantity x unit cost = subtotal) -----------

    def _build_resource_summary(self, wb: Workbook, estimate: EstimateResult, branding: BrandingConfig):
        ws = wb.create_sheet("Resources")
        accent = branding.primary_color_hex
        write_title(ws, "Selected Resources", accent, row=1)
        write_subtitle(
            ws,
            "Every resource on the estimate - Compute Engine, Persistent Disk, Cloud SQL, Networking, GKE, "
            "and every GCP service catalog selection - calculated separately. Resources with the same "
            "configuration are merged into one row with a combined quantity; different configurations never are.",
            row=2,
        )

        summaries = estimate.cost.resource_summaries
        header_row = 4
        headers = [
            "Category", "Resource", "Quantity", "Requested Configuration", "Normalized Configuration",
            "Unit Cost", "Subtotal", "Currency", "Region", "SKU", "Pricing Source", "Validation Status", "Assumption",
        ]
        write_header_row(ws, header_row, headers, accent)

        row = header_row + 1
        first_data_row = row
        if not summaries:
            write_data_row(ws, row, ["No resources on this estimate."] + [""] * (len(headers) - 1))
            row += 1
        else:
            for s in summaries:
                write_data_row(
                    ws, row,
                    [
                        s.category or "-", s.resource_name, s.quantity,
                        s.requested_configuration or s.configuration,
                        s.normalized_configuration or s.configuration,
                        s.unit_cost, s.subtotal, s.currency, s.region or "-",
                        s.sku_id or "(multiple)", s.pricing_source or "-",
                        s.status.replace("_", " ").title(), s.assumption_reason or "-",
                    ],
                    currency_cols={6, 7},
                )
                row += 1

        if summaries:
            total_row = row
            ws.cell(row=total_row, column=1, value="Grand Total").font = Font(bold=True)
            total_cell = ws.cell(row=total_row, column=7, value=f"=SUM(G{first_data_row}:G{row - 1})")
            total_cell.font = Font(bold=True)
            total_cell.number_format = "#,##0.00"
            for col in range(1, len(headers) + 1):
                ws.cell(row=total_row, column=col).fill = light_fill(accent)
            row = total_row + 1

        # -- Category Totals block: sum(category totals) == sum(resource
        # subtotals) == Grand Total above, and == the estimate's
        # subtotal_monthly (pre-discount/tax/support) - see
        # CostBreakdown.category_totals.
        row += 1
        cat_header_row = row
        write_header_row(ws, cat_header_row, ["Category", "Category Total"], accent)
        row += 1
        cat_first_row = row
        for category, total in estimate.cost.category_totals.items():
            write_data_row(ws, row, [category, total], currency_cols={2})
            row += 1
        if estimate.cost.category_totals:
            cat_total_row = row
            ws.cell(row=cat_total_row, column=1, value="Sum of Category Totals").font = Font(bold=True)
            cat_total_cell = ws.cell(
                row=cat_total_row, column=2,
                value=f"=SUM(B{cat_first_row}:B{row - 1})",
            )
            cat_total_cell.font = Font(bold=True)
            cat_total_cell.number_format = "#,##0.00"
            for col in (1, 2):
                ws.cell(row=cat_total_row, column=col).fill = light_fill(accent)

        freeze_header(ws, header_row)
        autofit_columns(ws, widths={
            1: 14, 2: 20, 3: 10, 4: 34, 5: 34, 6: 14, 7: 16, 8: 10, 9: 16, 10: 22, 11: 22, 12: 16, 13: 46,
        })
        ws.sheet_view.showGridLines = False

    # -- Licensing ------------------------------------------------------------

    def _build_licensing(self, wb: Workbook, estimate: EstimateResult, branding: BrandingConfig):
        ws = wb.create_sheet("Licensing")
        accent = branding.primary_color_hex
        write_title(ws, "Licensing", accent, row=1)
        write_subtitle(ws, "Software licensing model for each priced component.", row=2)

        header_row = 4
        write_header_row(ws, header_row, ["Component", "License Model", "Notes"], accent)
        row = header_row + 1

        engine = estimate.original_request.database.engine.value if (
            estimate.original_request.database and estimate.original_request.database.required
        ) else None

        os_value = estimate.normalized_spec.operating_system or "linux"
        os_license_notes = {
            "linux": ("Compute Engine (Linux)", "No additional license fee", "Standard public Linux images incur no separate OS license charge."),
            "windows_server": ("Compute Engine (Windows Server)", "Per-vCPU license fee", "Billed hourly per vCPU on top of the infrastructure price - see the Compute sheet's 'Licensing' line item."),
            "rhel": ("Compute Engine (RHEL)", "Flat per-instance license fee", "Red Hat Enterprise Linux subscription billed hourly per instance - see the Compute sheet's 'Licensing' line item."),
            "suse": ("Compute Engine (SUSE)", "Flat per-instance license fee", "SUSE Linux Enterprise Server subscription billed hourly per instance - see the Compute sheet's 'Licensing' line item."),
        }
        rows = [os_license_notes.get(os_value, os_license_notes["linux"])]
        if engine:
            license_note = (
                "License Included - Microsoft SQL Server license cost is bundled into the hourly Cloud SQL rate."
                if engine == "sqlserver"
                else "License Included - open-source engine; no separate license fee."
            )
            rows.append((f"Cloud SQL ({engine})", "License Included", license_note))
        if estimate.normalized_spec.kubernetes_node_count is not None:
            rows.append(("Google Kubernetes Engine", "No additional license fee", "GKE cluster management fee covers control plane; no separate software license."))

        for r in rows:
            write_data_row(ws, row, list(r))
            row += 1

        freeze_header(ws, header_row)
        autofit_columns(ws, widths={1: 30, 2: 26, 3: 60})
        ws.sheet_view.showGridLines = False

    # -- Assumptions ------------------------------------------------------------

    def _build_assumptions(
        self, wb: Workbook, estimate: EstimateResult, branding: BrandingConfig,
        explanation: EstimateExplanation | None = None,
    ):
        ws = wb.create_sheet("Assumptions")
        accent = branding.primary_color_hex
        write_title(ws, "Assumptions", accent, row=1)
        write_subtitle(ws, "Every substitution made while normalizing the request to a valid Google Cloud configuration.", row=2)

        has_ai_column = (
            explanation is not None
            and len(explanation.assumption_explanations) == len(estimate.assumptions)
        )
        headers = ["Field", "Requested Value", "Used Value", "Reason", "Strategy"]
        if has_ai_column:
            headers.append("Customer-Friendly Explanation")

        header_row = 4
        write_header_row(ws, header_row, headers, accent)
        row = header_row + 1
        if not estimate.assumptions:
            no_assumptions_row = ["-", "-", "-", "No assumptions were necessary; every requested value was directly supported.", "-"]
            if has_ai_column:
                no_assumptions_row.append("-")
            write_data_row(ws, row, no_assumptions_row)
            row += 1
        elif has_ai_column:
            for a, exp in zip(estimate.assumptions, explanation.assumption_explanations):
                write_data_row(ws, row, [a.field, a.requested_value, a.used_value, a.reason, a.strategy_applied or "-", exp.explanation])
                row += 1
        else:
            for a in estimate.assumptions:
                write_data_row(ws, row, [a.field, a.requested_value, a.used_value, a.reason, a.strategy_applied or "-"])
                row += 1

        freeze_header(ws, header_row)
        widths = {1: 24, 2: 24, 3: 30, 4: 60, 5: 16}
        if has_ai_column:
            widths[6] = 60
        autofit_columns(ws, widths=widths)
        ws.sheet_view.showGridLines = False

    # -- Validation ------------------------------------------------------------

    def _build_validation(self, wb: Workbook, estimate: EstimateResult, branding: BrandingConfig):
        ws = wb.create_sheet("Validation")
        accent = branding.primary_color_hex
        write_title(ws, "Validation Report", accent, row=1)
        write_subtitle(
            ws,
            f"{estimate.validation.blocker_count} blocker(s), {estimate.validation.warning_count} warning(s) found across {len(estimate.validation.results)} checks.",
            row=2,
        )

        header_row = 4
        write_header_row(ws, header_row, ["Field", "Rule", "Requested", "Supported", "Valid", "Severity", "Reason", "Recommendation"], accent)
        row = header_row + 1
        severity_colors = {"blocker": "F4C7C3", "warning": "FCE8B2", "info": "D9EAD3"}
        for r in estimate.validation.results:
            write_data_row(
                ws, row,
                [r.field, r.rule, r.requested_value, r.supported_value or "-", "Yes" if r.is_valid else "No",
                 r.severity.value.upper(), r.reason, r.recommendation],
            )
            fill_color = severity_colors.get(r.severity.value)
            if fill_color:
                from openpyxl.styles import PatternFill
                ws.cell(row=row, column=6).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            row += 1

        freeze_header(ws, header_row)
        autofit_columns(ws, widths={1: 22, 2: 22, 3: 16, 4: 16, 5: 8, 6: 12, 7: 50, 8: 50})
        ws.sheet_view.showGridLines = False

    # -- Recommendations (architecture) ------------------------------------------------------------

    def _build_recommendations(self, wb: Workbook, estimate: EstimateResult, branding: BrandingConfig):
        ws = wb.create_sheet("Recommendations")
        accent = branding.primary_color_hex
        write_title(ws, "Architecture Recommendations", accent, row=1)
        write_subtitle(ws, estimate.architecture.summary, row=2)

        header_row = 4
        write_header_row(ws, header_row, ["Layer", "Recommended Service", "Rationale"], accent)
        row = header_row + 1
        for c in estimate.architecture.components:
            write_data_row(ws, row, [c.layer, c.service, c.rationale])
            row += 1

        freeze_header(ws, header_row)
        autofit_columns(ws, widths={1: 20, 2: 40, 3: 60})
        ws.sheet_view.showGridLines = False

    # -- Pricing (full itemized) ------------------------------------------------------------

    def _build_pricing(self, wb: Workbook, estimate: EstimateResult, branding: BrandingConfig) -> int:
        ws = wb.create_sheet("Pricing")
        accent = branding.primary_color_hex
        write_title(ws, "Itemized Pricing", accent, row=1)
        write_subtitle(ws, f"All priced line items, sourced from: {estimate.cost.line_items[0].source if estimate.cost.line_items else 'n/a'}.", row=2)

        header_row = 4
        write_header_row(ws, header_row, ["Category", "Description", "SKU", "Unit", "Quantity", "Unit Price", "Monthly Amount", "Source"], accent)
        row = header_row + 1
        first_data_row = row
        for li in estimate.cost.line_items:
            write_data_row(
                ws, row,
                [li.category, li.description, li.sku_id or "-", li.unit, li.quantity, li.unit_price, li.monthly_amount, li.source],
                currency_cols={6, 7},
            )
            row += 1
        last_data_row = row - 1

        total_row = row
        ws.cell(row=total_row, column=1, value="Subtotal (Monthly)").font = Font(bold=True)
        if last_data_row >= first_data_row:
            total_cell = ws.cell(row=total_row, column=7, value=f"=SUM(G{first_data_row}:G{last_data_row})")
        else:
            total_cell = ws.cell(row=total_row, column=7, value=0)
        total_cell.font = Font(bold=True)
        total_cell.number_format = "#,##0.00"
        for col in range(1, 9):
            ws.cell(row=total_row, column=col).fill = light_fill(accent)

        freeze_header(ws, header_row)
        autofit_columns(ws, widths={1: 14, 2: 46, 3: 22, 4: 12, 5: 12, 6: 14, 7: 16, 8: 16})
        ws.sheet_view.showGridLines = False
        return total_row

    # -- Totals ------------------------------------------------------------

    def _build_totals(self, wb: Workbook, estimate: EstimateResult, branding: BrandingConfig, pricing_total_row: int):
        ws = wb.create_sheet("Totals")
        accent = branding.primary_color_hex
        write_title(ws, "Totals", accent, row=1)
        write_subtitle(ws, "Monthly cost roll-up, including discounts, tax, and support.", row=2)

        header_row = 4
        write_header_row(ws, header_row, ["Line", "Amount", "Currency"], accent)
        row = header_row + 1

        subtotal_row = row
        write_data_row(ws, row, ["Subtotal", f"=Pricing!G{pricing_total_row}", estimate.cost.currency])
        row += 1

        discount_rows_start = row
        if estimate.cost.discounts:
            for d in estimate.cost.discounts:
                write_data_row(ws, row, [f"Discount: {d.name} (-{d.percent_off}%)", -d.monthly_savings, estimate.cost.currency], currency_cols={2})
                row += 1
        else:
            write_data_row(ws, row, ["Discounts", 0, estimate.cost.currency], currency_cols={2})
            row += 1
        discount_rows_end = row - 1

        after_discount_row = row
        write_data_row(
            ws, row,
            [
                "After Discount",
                f"=B{subtotal_row}+SUM(B{discount_rows_start}:B{discount_rows_end})",
                estimate.cost.currency,
            ],
        )
        row += 1

        write_data_row(ws, row, [f"Tax ({estimate.cost.tax_rate_percent}%)", estimate.cost.tax_monthly, estimate.cost.currency], currency_cols={2})
        tax_row = row
        row += 1

        write_data_row(ws, row, [f"Support Plan ({estimate.cost.support_plan_percent}%)", estimate.cost.support_monthly, estimate.cost.currency], currency_cols={2})
        support_row = row
        row += 1

        total_row = row
        ws.cell(row=total_row, column=1, value="Total Monthly Cost").font = Font(bold=True, size=12)
        total_cell = ws.cell(row=total_row, column=2, value=f"=B{after_discount_row}+B{tax_row}+B{support_row}")
        total_cell.font = Font(bold=True, size=12)
        total_cell.number_format = "#,##0.00"
        ws.cell(row=total_row, column=3, value=estimate.cost.currency)
        for col in range(1, 4):
            ws.cell(row=total_row, column=col).fill = light_fill(accent)

        for r in (subtotal_row, after_discount_row):
            ws.cell(row=r, column=2).number_format = "#,##0.00"

        freeze_header(ws, header_row)
        autofit_columns(ws, widths={1: 30, 2: 20, 3: 12})
        ws.sheet_view.showGridLines = False

        self._totals_row_map = {
            "subtotal": subtotal_row,
            "after_discount": after_discount_row,
            "tax": tax_row,
            "support": support_row,
            "total": total_row,
        }

    # -- Yearly Cost ------------------------------------------------------------

    def _build_yearly(self, wb: Workbook, estimate: EstimateResult, branding: BrandingConfig):
        ws = wb.create_sheet("Yearly Cost")
        accent = branding.primary_color_hex
        write_title(ws, "Yearly Cost Projection", accent, row=1)
        write_subtitle(ws, "Monthly total extrapolated to yearly and 3-year horizons.", row=2)

        total_row = self._totals_row_map["total"]

        header_row = 4
        write_header_row(ws, header_row, ["Period", "Total Cost", "Currency"], accent)
        row = header_row + 1
        write_data_row(ws, row, ["Monthly", f"=Totals!B{total_row}", estimate.cost.currency])
        monthly_row = row
        row += 1
        write_data_row(ws, row, ["Yearly (x12)", f"=B{monthly_row}*12", estimate.cost.currency])
        row += 1
        write_data_row(ws, row, ["3-Year (x36)", f"=B{monthly_row}*36", estimate.cost.currency])
        row += 1

        row += 1
        write_header_row(ws, row, ["Month", "Cumulative Cost"], accent)
        row += 1
        month_header_row = row - 1
        for m in range(1, 13):
            write_data_row(ws, row, [f"Month {m}", f"=B{monthly_row}*{m}"])
            row += 1

        for r in range(header_row + 1, header_row + 4):
            ws.cell(row=r, column=2).number_format = "#,##0.00"
        for r in range(month_header_row + 1, row):
            ws.cell(row=r, column=2).number_format = "#,##0.00"

        freeze_header(ws, header_row)
        autofit_columns(ws, widths={1: 20, 2: 20, 3: 12})
        ws.sheet_view.showGridLines = False

    # -- Audit ------------------------------------------------------------

    def _build_audit(self, wb: Workbook, estimate: EstimateResult, branding: BrandingConfig):
        ws = wb.create_sheet("Audit")
        accent = branding.primary_color_hex
        write_title(ws, "Audit Log", accent, row=1)
        write_subtitle(ws, f"{len(estimate.audit_log.entries)} recorded decisions for estimate {estimate.estimate_id}.", row=2)

        header_row = 4
        write_header_row(ws, header_row, ["Timestamp (UTC)", "Actor", "Action", "Field", "Detail"], accent)
        row = header_row + 1
        for e in estimate.audit_log.entries:
            write_data_row(ws, row, [e.timestamp.strftime("%Y-%m-%d %H:%M:%S"), e.actor, e.action, e.field or "-", e.detail])
            row += 1

        freeze_header(ws, header_row)
        autofit_columns(ws, widths={1: 20, 2: 16, 3: 22, 4: 24, 5: 70})
        ws.sheet_view.showGridLines = False
