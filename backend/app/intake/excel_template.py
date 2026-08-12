"""
Generates the blank Excel questionnaire customers/consultants fill in and
upload back through `POST /api/v1/intake/excel`. Driven entirely by
`FIELD_SCHEMA` (see `schema.py`) so it can never drift from what the parser
actually reads.
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from app.intake.schema import FIELD_SCHEMA, SECTION_ORDER, TOP_LEVEL_SECTION
from app.reports.styles import autofit_columns, write_header_row, write_subtitle, write_title

ACCENT = "1A73E8"


class ExcelTemplateGenerator:
    def generate(self) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Questionnaire"

        write_title(ws, "Cloud Infrastructure Requirements Questionnaire", ACCENT, row=1, span=3)
        write_subtitle(
            ws,
            "Fill in the Value column only. Leave an entire section blank if it doesn't apply - for example, "
            "leave Compute blank and fill in Business Context instead if you only know approximate user counts.",
            row=2, span=3,
        )

        header_row = 4
        write_header_row(ws, header_row, ["Field", "Value", "Example / Notes"], ACCENT)
        row = header_row + 1

        by_section: dict[str, list] = {}
        for spec in FIELD_SCHEMA:
            by_section.setdefault(spec.section, []).append(spec)

        for section in SECTION_ORDER:
            specs = by_section.get(section, [])
            if not specs:
                continue

            if section != TOP_LEVEL_SECTION:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
                section_cell = ws.cell(row=row, column=1, value=section)
                section_cell.font = Font(bold=True, color="FFFFFF")
                fill = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type="solid")
                for col in range(1, 4):
                    ws.cell(row=row, column=col).fill = fill
                row += 1

            for spec in specs:
                label_cell = ws.cell(row=row, column=1, value=spec.label)
                label_cell.font = Font(bold=True)
                ws.cell(row=row, column=2, value=None)
                example_cell = ws.cell(row=row, column=3, value=spec.example)
                example_cell.font = Font(italic=True, size=9, color="888888")

                if spec.choices:
                    dv = DataValidation(
                        type="list", formula1='"' + ",".join(spec.choices) + '"',
                        allow_blank=True, showDropDown=False,
                    )
                    ws.add_data_validation(dv)
                    dv.add(ws.cell(row=row, column=2))
                elif spec.value_type == "bool":
                    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True, showDropDown=False)
                    ws.add_data_validation(dv)
                    dv.add(ws.cell(row=row, column=2))

                row += 1

        autofit_columns(ws, widths={1: 30, 2: 22, 3: 55})
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
