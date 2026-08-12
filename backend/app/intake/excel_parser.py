"""
Parses a filled-in questionnaire workbook (see `excel_template.py`) into a
`CustomerRequirement`.

Design principle carried over from the validation/normalization engines:
never trust the input, and never let one bad cell blow up the whole parse.
A typo in the Database section still lets the Compute section come through
cleanly - the caller sees exactly which section/field failed and why via
`ParseIssue`, the same "always explain, never hide" philosophy as
`Assumption` and `ValidationResult`.
"""
import io

import openpyxl
from pydantic import ValidationError

from app.core.exceptions import IntakeParseError
from app.domain.enums import Severity
from app.domain.intake import ParseIssue
from app.domain.requirements import (
    AvailabilityRequirement,
    BusinessContext,
    ComputeRequirement,
    CustomerRequirement,
    DatabaseRequirement,
    KubernetesRequirement,
    NetworkRequirement,
    StorageRequirement,
)
from app.intake.schema import FIELD_SCHEMA, SECTION_KEYS

_SECTION_MODELS = {
    "compute": ComputeRequirement,
    "storage": StorageRequirement,
    "database": DatabaseRequirement,
    "network": NetworkRequirement,
    "kubernetes": KubernetesRequirement,
    "availability": AvailabilityRequirement,
    "business": BusinessContext,
}

_TRUE_STRINGS = {"true", "yes", "y", "1"}
_FALSE_STRINGS = {"false", "no", "n", "0"}


class ExcelQuestionnaireParser:
    def parse(self, file_bytes: bytes) -> tuple[CustomerRequirement | None, list[ParseIssue]]:
        raw_values = self._read_cells(file_bytes)
        issues: list[ParseIssue] = []

        top_level: dict = {}
        sections: dict[str, dict] = {key: {} for key in SECTION_KEYS.values()}

        label_lookup = {spec.label.strip().lower(): spec for spec in FIELD_SCHEMA}

        for label, raw_value in raw_values.items():
            spec = label_lookup.get(label)
            if spec is None or raw_value is None or (isinstance(raw_value, str) and raw_value.strip() == ""):
                continue

            try:
                coerced = self._coerce(raw_value, spec)
            except (TypeError, ValueError):
                issues.append(ParseIssue(
                    field=spec.label,
                    message=f"Could not interpret '{raw_value}' as a {spec.value_type} value; this field was skipped.",
                    severity=Severity.WARNING,
                ))
                continue

            if len(spec.path) == 1:
                top_level[spec.path[0]] = coerced
            else:
                section_key, field_key = spec.path
                sections[section_key][field_key] = coerced

        if not top_level.get("project_name"):
            issues.append(ParseIssue(
                field="Project Name", message="Project Name is required but was blank.", severity=Severity.BLOCKER,
            ))
            return None, issues

        built_sections: dict = {}
        for section_key, values in sections.items():
            if not values:
                built_sections[section_key] = None
                continue
            model_cls = _SECTION_MODELS[section_key]
            try:
                built_sections[section_key] = model_cls(**values)
            except ValidationError as exc:
                section_label = next(k for k, v in SECTION_KEYS.items() if v == section_key)
                for err in exc.errors():
                    field_name = ".".join(str(p) for p in err["loc"]) or section_key
                    issues.append(ParseIssue(
                        field=f"{section_label}.{field_name}",
                        message=f"{err['msg']} - the entire {section_label} section was skipped.",
                        severity=Severity.BLOCKER,
                    ))
                built_sections[section_key] = None

        try:
            requirement = CustomerRequirement(**top_level, **built_sections)
        except ValidationError as exc:
            for err in exc.errors():
                issues.append(ParseIssue(
                    field=".".join(str(p) for p in err["loc"]), message=err["msg"], severity=Severity.BLOCKER,
                ))
            return None, issues

        return requirement, issues

    def _read_cells(self, file_bytes: bytes) -> dict[str, object]:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        except Exception as exc:  # openpyxl raises several distinct types for corrupt files
            raise IntakeParseError(
                "The uploaded file is not a valid Excel (.xlsx) workbook, or could not be opened."
            ) from exc

        ws = None
        for name in wb.sheetnames:
            if name.strip().lower() == "questionnaire":
                ws = wb[name]
                break
        if ws is None:
            ws = wb[wb.sheetnames[0]]

        values: dict[str, object] = {}
        for row in ws.iter_rows():
            if len(row) < 2:
                continue
            label_cell, value_cell = row[0], row[1]
            if label_cell.value is None:
                continue
            label = str(label_cell.value).strip().lower()
            values[label] = value_cell.value
        return values

    def _coerce(self, raw_value: object, spec) -> object:
        if spec.value_type == "str":
            return str(raw_value).strip()

        if spec.value_type == "int":
            if isinstance(raw_value, bool):
                raise ValueError("boolean given for an integer field")
            return int(float(raw_value)) if isinstance(raw_value, str) else int(raw_value)

        if spec.value_type == "float":
            if isinstance(raw_value, bool):
                raise ValueError("boolean given for a float field")
            return float(raw_value)

        if spec.value_type == "bool":
            if isinstance(raw_value, bool):
                return raw_value
            text = str(raw_value).strip().lower()
            if text in _TRUE_STRINGS:
                return True
            if text in _FALSE_STRINGS:
                return False
            raise ValueError(f"'{raw_value}' is not a recognized boolean")

        if spec.value_type == "enum":
            text = str(raw_value).strip()
            for choice in spec.choices or []:
                if choice.lower() == text.lower():
                    return choice
            raise ValueError(f"'{raw_value}' is not one of {spec.choices}")

        raise ValueError(f"Unknown field type '{spec.value_type}'")
