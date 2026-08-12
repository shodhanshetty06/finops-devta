"""
Full user-journey integration test: one continuous session walking through
most of the platform against a real HTTP server, proving the pieces work
together end to end - not just each endpoint in isolation (that's what
`backend/tests/` already covers exhaustively). See tests_integration/README.md.
"""
import io
import time

import httpx
import pytest
from openpyxl import load_workbook

COMPUTE_REQUIREMENT = {
    "project_name": "Integration Journey",
    "region": "us-central1",
    "normalization_strategy": "balanced",
    "compute": {"machine_family": "n2", "vcpu": 4, "ram_gb": 16, "instance_count": 2},
    "storage": {"disk_type": "pd-balanced", "size_gb": 200, "snapshot_enabled": True},
    "database": {"required": True, "engine": "postgres", "size_gb": 50, "vcpu": 2, "ram_gb": 8},
    "network": {"load_balancer_required": True, "external_ip_count": 1, "estimated_egress_gb_per_month": 300},
}


def _run_full_journey(base_url: str) -> None:
    # trust_env=False: this sandbox sets a SOCKS proxy for external network
    # egress allowlisting - it must never be applied to our own localhost
    # subprocess (see the matching note in conftest.py::_wait_for_health).
    client = httpx.Client(base_url=base_url, timeout=30.0, trust_env=False)

    # -- Phase 3: register + login ------------------------------------------
    email = f"integration-{int(time.time() * 1000)}@example.com"
    resp = client.post("/api/v1/auth/register", json={
        "email": email, "password": "supersecret123", "full_name": "Integration Test", "role": "customer",
    })
    assert resp.status_code == 201, resp.text

    resp = client.post("/api/v1/auth/login", data={"username": email, "password": "supersecret123"})
    assert resp.status_code == 200, resp.text
    token = resp.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    resp = client.get("/api/v1/auth/me", headers=headers)
    assert resp.status_code == 200
    assert resp.json()["email"] == email

    # -- Phase 4: download template, fill it in, upload as intake -----------
    resp = client.get("/api/v1/intake/excel/template")
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb["Questionnaire"]
    values = {
        "Project Name": "Integration Journey (from Excel)", "Region": "us-central1",
        "Machine Family": "n2", "vCPU": 4, "RAM (GB)": 16, "Instance Count": 1,
    }
    for row in ws.iter_rows():
        if row[0].value in values:
            row[1].value = values[row[0].value]
    buf = io.BytesIO()
    wb.save(buf)

    resp = client.post(
        "/api/v1/intake/excel",
        files={"file": ("q.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        params={"auto_estimate": "true"},
    )
    assert resp.status_code == 200, resp.text
    assert resp.json()["estimate"] is not None

    # -- Phase 3: create a project, save the first estimate version ---------
    resp = client.post("/api/v1/projects", json={"name": "Integration Project"}, headers=headers)
    assert resp.status_code == 201, resp.text
    project = resp.json()
    assert project["monthly_budget_usd"] is None  # Phase 10 -> Phase 8 field, default unset

    resp = client.post(
        f"/api/v1/projects/{project['id']}/estimates",
        json={"requirement": COMPUTE_REQUIREMENT}, headers=headers,
    )
    assert resp.status_code == 201, resp.text
    v1 = resp.json()
    assert v1["version"] == 1
    actual_monthly = v1["result"]["cost"]["total_monthly"]
    assert actual_monthly > 0
    assert v1["budget_status"]["budget_monthly"] is None

    # -- Phase 8: optimization engines, all reusing the real pricing pipeline --
    resp = client.post(
        "/api/v1/optimization/rightsizing", headers=headers,
        json={
            "requirement": COMPUTE_REQUIREMENT,
            "usage": {"avg_cpu_utilization_percent": 12, "peak_cpu_utilization_percent": 18, "observation_period_days": 30},
        },
    )
    assert resp.status_code == 200, resp.text
    assert resp.json()["findings"][0]["action"] == "downsize"

    resp = client.post(
        "/api/v1/optimization/commitment-recommendation", headers=headers,
        json={"requirement": COMPUTE_REQUIREMENT, "workload_stability": "steady"},
    )
    assert resp.status_code == 200, resp.text
    assert resp.json()["on_demand_discountable_monthly_cost"] > 0

    resp = client.post(
        "/api/v1/optimization/forecast", headers=headers,
        json={"requirement": COMPUTE_REQUIREMENT, "monthly_growth_percent": 4, "months": 6},
    )
    assert resp.status_code == 200, resp.text
    assert len(resp.json()["points"]) == 6

    resp = client.post("/api/v1/optimization/carbon", headers=headers, json={"requirement": COMPUTE_REQUIREMENT})
    assert resp.status_code == 200, resp.text
    assert resp.json()["estimated_kgco2e_per_month"] > 0

    resp = client.post(
        "/api/v1/optimization/compare-regions", headers=headers,
        json={"requirement": COMPUTE_REQUIREMENT, "regions": ["us-central1", "europe-west1", "asia-south1"]},
    )
    assert resp.status_code == 200, resp.text
    assert len(resp.json()["options"]) == 3

    # -- Phase 9: cross-cloud comparison -------------------------------------
    resp = client.post(
        "/api/v1/optimization/compare-clouds", headers=headers,
        json={"requirement": COMPUTE_REQUIREMENT},
    )
    assert resp.status_code == 200, resp.text
    cloud_options = {o["cloud_provider"] for o in resp.json()["options"]}
    assert cloud_options == {"gcp", "aws", "azure"}

    # -- Phase 8: budget tracking + a second (bigger) version + comparison --
    resp = client.patch(
        f"/api/v1/projects/{project['id']}/budget", headers=headers,
        json={"monthly_budget_usd": actual_monthly / 2},
    )
    assert resp.status_code == 200, resp.text

    bigger_requirement = dict(COMPUTE_REQUIREMENT, compute={
        "machine_family": "n2", "vcpu": 16, "ram_gb": 64, "instance_count": 2,
    })
    resp = client.post(
        f"/api/v1/projects/{project['id']}/estimates", headers=headers,
        json={"requirement": bigger_requirement},
    )
    assert resp.status_code == 201, resp.text
    v2 = resp.json()
    assert v2["version"] == 2
    assert v2["budget_status"]["within_budget"] is False  # budget was set to half of v1's cost

    resp = client.get(f"/api/v1/projects/{project['id']}/estimates/compare?from=1&to=2", headers=headers)
    assert resp.status_code == 200, resp.text
    comparison = resp.json()
    assert comparison["delta_monthly"] > 0
    assert comparison["to_total_monthly"] > comparison["from_total_monthly"]

    # -- Phase 2: export the saved version as Excel and PDF -----------------
    resp = client.post(f"/api/v1/projects/{project['id']}/estimates/2/reports/excel", headers=headers, json={})
    assert resp.status_code == 200
    assert resp.content[:2] == b"PK"

    resp = client.post(f"/api/v1/projects/{project['id']}/estimates/2/reports/pdf", headers=headers, json={})
    assert resp.status_code == 200
    assert resp.content[:5] == b"%PDF-"

    # -- Phase 7: async report job, enqueued and polled to completion -------
    resp = client.post(
        "/api/v1/jobs/reports",
        json={"estimate": v2["result"], "format": "excel"},
    )
    assert resp.status_code == 202, resp.text
    job_id = resp.json()["job_id"]

    deadline = time.time() + 15
    job_status = None
    while time.time() < deadline:
        resp = client.get(f"/api/v1/jobs/{job_id}")
        assert resp.status_code == 200
        job_status = resp.json()
        if job_status["ready"]:
            break
        time.sleep(0.2)
    assert job_status is not None and job_status["ready"], f"job never completed: {job_status}"
    assert job_status["successful"] is True

    resp = client.get(f"/api/v1/jobs/{job_id}/download")
    assert resp.status_code == 200
    assert resp.content[:2] == b"PK"

    # -- RBAC: a second user cannot see the first user's project -------------
    other_email = f"integration-other-{int(time.time() * 1000)}@example.com"
    client.post("/api/v1/auth/register", json={
        "email": other_email, "password": "supersecret123", "full_name": "Other User", "role": "customer",
    })
    resp = client.post("/api/v1/auth/login", data={"username": other_email, "password": "supersecret123"})
    other_headers = {"Authorization": f"Bearer {resp.json()['access_token']}"}

    resp = client.get(f"/api/v1/projects/{project['id']}", headers=other_headers)
    assert resp.status_code == 403


def test_full_user_journey_sqlite(live_server):
    _run_full_journey(live_server)


def test_full_user_journey_postgres(live_server_postgres):
    """Same journey, against the real Postgres docker-compose.yml deploys -
    skipped unless FINOPS_TEST_POSTGRES_URL is set (see README.md). Exists
    to catch SQLite-vs-Postgres dialect differences (e.g. JSON column
    behavior, case sensitivity) that the main suite's SQLite-only tests
    structurally cannot catch."""
    _run_full_journey(live_server_postgres)
