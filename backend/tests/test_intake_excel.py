"""Tests for the Excel questionnaire template + parser (Phase 4)."""
import io

from openpyxl import load_workbook

from app.intake.excel_parser import ExcelQuestionnaireParser
from app.intake.excel_template import ExcelTemplateGenerator
from app.intake.schema import FIELD_SCHEMA


def _blank_workbook():
    blank_bytes = ExcelTemplateGenerator().generate()
    return load_workbook(io.BytesIO(blank_bytes))


def _fill_and_save(wb, values: dict[str, object]) -> bytes:
    ws = wb["Questionnaire"]
    for row in ws.iter_rows():
        label = row[0].value
        if label in values:
            row[1].value = values[label]
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_field_schema_has_no_duplicate_labels():
    labels = [spec.label.strip().lower() for spec in FIELD_SCHEMA]
    assert len(labels) == len(set(labels)), "every questionnaire label must be globally unique for the parser to work"


def test_template_generates_valid_workbook_with_questionnaire_sheet():
    wb = _blank_workbook()
    assert "Questionnaire" in wb.sheetnames
    ws = wb["Questionnaire"]
    labels = {row[0].value for row in ws.iter_rows() if row[0].value}
    for spec in FIELD_SCHEMA:
        assert spec.label in labels


def test_full_round_trip_all_sections_populated():
    wb = _blank_workbook()
    filled = _fill_and_save(wb, {
        "Project Name": "Retail Inventory API",
        "Region": "us-central1",
        "Machine Family": "e2",
        "vCPU": 4,
        "RAM (GB)": 16,
        "Instance Count": 2,
        "Disk Type": "pd-balanced",
        "Size (GB)": 157,
        "Snapshot Enabled": "TRUE",
        "Snapshot Retention (Days)": 14,
        "Database Required": "TRUE",
        "Engine": "postgres",
        "Database Size (GB)": 100,
        "Database vCPU": 2,
        "Database RAM (GB)": 8,
        "Database High Availability": "TRUE",
        "Load Balancer Required": "TRUE",
        "External IP Count": 1,
        "Kubernetes Required": "FALSE",
        "High Availability": "TRUE",
        "Target Uptime (%)": 99.95,
        "Backup Frequency": "daily",
    })

    req, issues = ExcelQuestionnaireParser().parse(filled)

    assert issues == []
    assert req.project_name == "Retail Inventory API"
    assert req.compute.vcpu == 4 and req.compute.ram_gb == 16
    assert req.storage.size_gb == 157 and req.storage.snapshot_enabled is True
    assert req.database.required is True and req.database.high_availability is True
    assert req.network.load_balancer_required is True
    assert req.availability.high_availability is True
    assert req.availability.target_uptime_percent == 99.95


def test_partial_section_is_dropped_with_blocker_but_other_sections_survive():
    wb = _blank_workbook()
    filled = _fill_and_save(wb, {
        "Project Name": "Partial Sheet",
        "vCPU": 4,
        "RAM (GB)": 16,
        # Storage section only partially filled - size_gb (required) is missing.
        "Snapshot Enabled": "TRUE",
    })

    req, issues = ExcelQuestionnaireParser().parse(filled)

    assert req is not None
    assert req.compute is not None and req.compute.vcpu == 4
    assert req.storage is None
    blocker_fields = [i.field for i in issues if i.severity.value == "blocker"]
    assert "Storage.size_gb" in blocker_fields


def test_invalid_cell_values_are_skipped_not_fatal():
    wb = _blank_workbook()
    filled = _fill_and_save(wb, {
        "Project Name": "Bad Cells",
        "vCPU": "not-a-number",
        "RAM (GB)": 16,
        "Machine Family": "does-not-exist",
    })

    req, issues = ExcelQuestionnaireParser().parse(filled)

    assert req is not None  # project still parses
    assert req.compute is None  # vcpu never made it in, so the section can't be built
    warnings = [i for i in issues if i.severity.value == "warning"]
    assert any("vCPU" in w.field for w in warnings)
    assert any("Machine Family" in w.field for w in warnings)


def test_missing_project_name_gets_an_auto_default_with_a_warning():
    wb = _blank_workbook()
    filled = _fill_and_save(wb, {"vCPU": 4, "RAM (GB)": 16})

    req, issues = ExcelQuestionnaireParser().parse(filled)

    assert req is not None  # a blank Project Name no longer blocks parsing/pricing
    assert req.project_name.startswith("Quick Estimate")
    assert req.compute is not None and req.compute.vcpu == 4
    assert any(i.severity.value == "warning" and "Project Name" in i.field for i in issues)
    assert not any(i.severity.value == "blocker" and "Project Name" in i.field for i in issues)


def test_unreadable_file_raises_intake_parse_error():
    from app.core.exceptions import IntakeParseError

    parser = ExcelQuestionnaireParser()
    try:
        parser.parse(b"this is not a valid xlsx file at all")
        assert False, "expected IntakeParseError"
    except IntakeParseError:
        pass


def test_blank_workbook_produces_default_named_requirement_and_no_crash():
    blank_bytes = ExcelTemplateGenerator().generate()
    req, issues = ExcelQuestionnaireParser().parse(blank_bytes)
    assert req is not None
    assert req.project_name.startswith("Quick Estimate")
    assert not any(i.severity.value == "blocker" for i in issues)
