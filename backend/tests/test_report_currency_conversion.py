"""
Tests for report-export currency conversion:
- CurrencyConverter.convert_estimate (app/pricing/currency_converter.py)
- convert_estimate_currency's never-raises fallback wrapper
- POST /api/v1/reports/{excel,pdf} target_currency
- POST /api/v1/projects/{id}/estimates/{version}/reports/{excel,pdf}?target_currency=...

No real network call is made anywhere - httpx.MockTransport intercepts every
request to the Frankfurter API, same pattern as tests/test_currency.py.
"""
import io

import httpx
import pytest
from openpyxl import load_workbook

from app.core.exceptions import CurrencyProviderError
from app.domain.branding import BrandingConfig
from app.main import app
from app.pricing.currency_converter import (
    CurrencyConverter,
    CurrencyExchangeClient,
    InMemoryCurrencyRateCache,
    convert_estimate_currency,
    get_currency_converter,
)
from tests.conftest import register_and_login

pypdf = pytest.importorskip("pypdf")


def _frankfurter_response(base="USD", rates=None, date="2026-08-08"):
    rates = rates or {"INR": 80.0, "EUR": 0.9}
    return httpx.Response(200, json={"amount": 1.0, "base": base, "date": date, "rates": rates})


def _converter(handler):
    client = CurrencyExchangeClient("https://api.frankfurter.dev/v1/latest", transport=httpx.MockTransport(handler))
    return CurrencyConverter(client, InMemoryCurrencyRateCache(), ttl_seconds=60, symbols=["INR", "EUR"])


# -- CurrencyConverter.convert_estimate ---------------------------------------

def test_convert_estimate_returns_same_object_when_already_in_target_currency(sample_estimate):
    converter = _converter(lambda r: _frankfurter_response())
    result = converter.convert_estimate(sample_estimate, sample_estimate.cost.currency)
    assert result is sample_estimate  # no rate fetch, no copy - identity fast-path


def test_convert_estimate_rescales_every_monetary_figure(sample_estimate):
    converter = _converter(lambda r: _frankfurter_response(rates={"INR": 80.0}))
    converted = converter.convert_estimate(sample_estimate, "INR")

    assert converted.cost.currency == "INR"
    assert converted.cost.total_monthly == round(sample_estimate.cost.total_monthly * 80.0, 2)
    assert converted.cost.total_yearly == round(sample_estimate.cost.total_yearly * 80.0, 2)
    assert converted.cost.subtotal_monthly == round(sample_estimate.cost.subtotal_monthly * 80.0, 2)

    # Percentages are currency-independent and must be untouched.
    assert converted.cost.tax_rate_percent == sample_estimate.cost.tax_rate_percent
    assert converted.cost.support_plan_percent == sample_estimate.cost.support_plan_percent

    # Every line item and resource summary is rescaled and re-tagged.
    assert len(converted.cost.line_items) == len(sample_estimate.cost.line_items)
    for original, conv in zip(sample_estimate.cost.line_items, converted.cost.line_items):
        assert conv.currency == "INR"
        assert conv.monthly_amount == round(original.monthly_amount * 80.0, 2)

    assert len(converted.cost.resource_summaries) == len(sample_estimate.cost.resource_summaries)
    for original, conv in zip(sample_estimate.cost.resource_summaries, converted.cost.resource_summaries):
        assert conv.currency == "INR"
        assert conv.subtotal == round(original.subtotal * 80.0, 2)

    for category, total in sample_estimate.cost.category_totals.items():
        assert converted.cost.category_totals[category] == round(total * 80.0, 2)

    # The original estimate object must never be mutated.
    assert sample_estimate.cost.currency != "INR"


def test_convert_estimate_raises_when_no_rate_available(sample_estimate):
    converter = _converter(lambda r: _frankfurter_response(rates={"EUR": 0.9}))  # no INR
    with pytest.raises(CurrencyProviderError):
        converter.convert_estimate(sample_estimate, "INR")


# -- convert_estimate_currency (never-raises wrapper) --------------------------

def test_convert_estimate_currency_noop_when_target_is_none(sample_estimate):
    converter = _converter(lambda r: _frankfurter_response())
    result = convert_estimate_currency(sample_estimate, None, converter)
    assert result is sample_estimate


def test_convert_estimate_currency_falls_back_to_native_currency_on_failure(sample_estimate):
    converter = _converter(lambda r: httpx.Response(503, text="down"))
    result = convert_estimate_currency(sample_estimate, "INR", converter)
    # Must never raise, and must not silently claim to be in INR when the
    # conversion actually failed - falls back to the original, untouched.
    assert result is sample_estimate
    assert result.cost.currency != "INR"


# -- API: POST /api/v1/reports/{excel,pdf} target_currency ---------------------
# `api_client` (isolated per-test SQLite db) comes from tests/conftest.py -
# not redefined here, so it stays consistent with test_projects.py's setup.

@pytest.fixture(autouse=True)
def _stub_currency_converter():
    converter = _converter(lambda r: _frankfurter_response(rates={"INR": 80.0, "EUR": 0.9}))
    app.dependency_overrides[get_currency_converter] = lambda: converter
    yield
    app.dependency_overrides.clear()


def test_reports_pdf_endpoint_renders_in_target_currency(api_client, sample_estimate):
    resp = api_client.post("/api/v1/reports/pdf", json={
        "estimate": sample_estimate.model_dump(mode="json"),
        "branding": BrandingConfig().model_dump(mode="json"),
        "target_currency": "INR",
    })
    assert resp.status_code == 200
    reader = pypdf.PdfReader(io.BytesIO(resp.content))
    text = "\n".join(p.extract_text() for p in reader.pages)
    expected_total = round(sample_estimate.cost.total_monthly * 80.0, 2)
    assert f"INR {expected_total:,.2f}" in text
    assert "USD" not in text


def test_reports_excel_endpoint_renders_in_target_currency(api_client, sample_estimate):
    resp = api_client.post("/api/v1/reports/excel", json={
        "estimate": sample_estimate.model_dump(mode="json"),
        "branding": BrandingConfig().model_dump(mode="json"),
        "target_currency": "INR",
    })
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb["Summary"]
    values = [cell.value for row in ws.iter_rows() for cell in row]
    assert "INR" in values
    assert round(sample_estimate.cost.total_monthly * 80.0, 2) in values


def test_reports_pdf_endpoint_without_target_currency_stays_native(api_client, sample_estimate):
    resp = api_client.post("/api/v1/reports/pdf", json={
        "estimate": sample_estimate.model_dump(mode="json"),
        "branding": BrandingConfig().model_dump(mode="json"),
    })
    assert resp.status_code == 200
    reader = pypdf.PdfReader(io.BytesIO(resp.content))
    text = "\n".join(p.extract_text() for p in reader.pages)
    assert f"{sample_estimate.cost.currency} {sample_estimate.cost.total_monthly:,.2f}" in text


# -- API: saved-version export with ?target_currency= query param -------------

def test_saved_version_excel_export_applies_target_currency_query_param(api_client, sample_estimate):
    _, headers = register_and_login(api_client, "currency-export-test@example.com")
    project_resp = api_client.post("/api/v1/projects", json={"name": "Currency Export Test"}, headers=headers)
    assert project_resp.status_code == 201
    project_id = project_resp.json()["id"]

    est_resp = api_client.post(
        f"/api/v1/projects/{project_id}/estimates",
        json={"requirement": sample_estimate.original_request.model_dump(mode="json")},
        headers=headers,
    )
    assert est_resp.status_code == 201, est_resp.text
    saved_currency = est_resp.json()["result"]["cost"]["currency"]
    saved_total_monthly = est_resp.json()["result"]["cost"]["total_monthly"]

    resp = api_client.post(
        f"/api/v1/projects/{project_id}/estimates/1/reports/excel",
        params={"target_currency": "INR"},
        json={},
        headers=headers,
    )
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb["Summary"]
    values = [cell.value for row in ws.iter_rows() for cell in row]
    assert "INR" in values
    if saved_currency != "INR":
        assert round(saved_total_monthly * 80.0, 2) in values


def test_saved_version_excel_export_without_query_param_stays_native(api_client, sample_estimate):
    _, headers = register_and_login(api_client, "currency-export-native-test@example.com")
    project_resp = api_client.post("/api/v1/projects", json={"name": "Currency Export Native Test"}, headers=headers)
    project_id = project_resp.json()["id"]

    est_resp = api_client.post(
        f"/api/v1/projects/{project_id}/estimates",
        json={"requirement": sample_estimate.original_request.model_dump(mode="json")},
        headers=headers,
    )
    assert est_resp.status_code == 201, est_resp.text
    saved_currency = est_resp.json()["result"]["cost"]["currency"]

    resp = api_client.post(
        f"/api/v1/projects/{project_id}/estimates/1/reports/excel",
        json={},
        headers=headers,
    )
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb["Summary"]
    values = [cell.value for row in ws.iter_rows() for cell in row]
    assert saved_currency in values
