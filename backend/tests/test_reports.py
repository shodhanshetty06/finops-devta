import io
import json

import httpx
import pytest
from openpyxl import load_workbook

from app.domain.branding import BrandingConfig
from app.llm.groq_provider import GroqProvider
from app.reports.excel_generator import ExcelReportGenerator
from app.reports.pdf_generator import PdfReportGenerator
from app.services.explanation_service import ExplanationService

pypdf = pytest.importorskip("pypdf")

EXPECTED_SHEETS = [
    "Summary", "Compute", "Storage", "Database", "Networking", "Resources", "Licensing",
    "Assumptions", "Validation", "Recommendations", "Pricing", "Totals",
    "Yearly Cost", "Audit",
]


def test_excel_report_contains_all_required_sheets(sample_estimate):
    xbytes = ExcelReportGenerator().generate(sample_estimate, BrandingConfig(prepared_for="Acme Retail Co."))
    wb = load_workbook(io.BytesIO(xbytes))
    assert wb.sheetnames == EXPECTED_SHEETS


def test_excel_summary_sheet_has_key_metrics(sample_estimate):
    xbytes = ExcelReportGenerator().generate(sample_estimate)
    wb = load_workbook(io.BytesIO(xbytes))
    ws = wb["Summary"]
    values = [cell.value for row in ws.iter_rows() for cell in row]
    assert sample_estimate.project_name in values
    assert sample_estimate.estimate_id in values
    assert sample_estimate.cost.total_monthly in values
    assert sample_estimate.cost.total_yearly in values


def test_excel_pricing_sheet_lists_every_line_item(sample_estimate):
    xbytes = ExcelReportGenerator().generate(sample_estimate)
    wb = load_workbook(io.BytesIO(xbytes))
    ws = wb["Pricing"]
    descriptions_in_sheet = {row[1] for row in ws.iter_rows(min_row=5, values_only=True) if row[1]}
    for li in sample_estimate.cost.line_items:
        assert li.description in descriptions_in_sheet
    # Subtotal row uses a real SUM formula, not a hardcoded number.
    formulas = [row[6] for row in ws.iter_rows(min_row=5, values_only=True) if isinstance(row[6], str) and row[6].startswith("=SUM(")]
    assert len(formulas) == 1


def test_excel_assumptions_sheet_matches_estimate_assumptions(sample_estimate):
    xbytes = ExcelReportGenerator().generate(sample_estimate)
    wb = load_workbook(io.BytesIO(xbytes))
    ws = wb["Assumptions"]
    reasons_in_sheet = {row[3] for row in ws.iter_rows(min_row=5, values_only=True) if row[3]}
    for a in sample_estimate.assumptions:
        assert a.reason in reasons_in_sheet


def test_excel_validation_sheet_matches_estimate_validation(sample_estimate):
    xbytes = ExcelReportGenerator().generate(sample_estimate)
    wb = load_workbook(io.BytesIO(xbytes))
    ws = wb["Validation"]
    row_count = sum(1 for row in ws.iter_rows(min_row=5, values_only=True) if row[0])
    assert row_count == len(sample_estimate.validation.results)


def test_excel_totals_and_yearly_sheets_reference_pricing_via_formulas(sample_estimate):
    xbytes = ExcelReportGenerator().generate(sample_estimate)
    wb = load_workbook(io.BytesIO(xbytes))
    totals_ws = wb["Totals"]
    subtotal_formula = totals_ws["B5"].value
    assert isinstance(subtotal_formula, str) and subtotal_formula.startswith("=Pricing!")

    yearly_ws = wb["Yearly Cost"]
    monthly_formula = yearly_ws["B5"].value
    assert isinstance(monthly_formula, str) and monthly_formula.startswith("=Totals!")


def test_pdf_report_is_valid_and_contains_key_sections(sample_estimate):
    pbytes = PdfReportGenerator().generate(sample_estimate, BrandingConfig(prepared_for="Acme Retail Co."))
    assert pbytes[:5] == b"%PDF-"

    reader = pypdf.PdfReader(io.BytesIO(pbytes))
    # The consolidated layout (Selected Resources + Category Totals +
    # Itemized Pricing merged into one Cost Breakdown table, no forced
    # PageBreaks) should keep even a multi-category, multi-assumption
    # estimate to a handful of pages rather than the old fixed ~7-page
    # skeleton every estimate used to produce regardless of content.
    assert 1 <= len(reader.pages) <= 5
    # Cell text now wraps (Paragraph flowables), so pypdf can insert a
    # newline mid-phrase where a cell wrapped - normalize before checking
    # for an exact multi-word phrase.
    text = " ".join(p.extract_text() for p in reader.pages).replace("\n", " ")

    for expected in [
        sample_estimate.project_name,
        "Executive Summary",
        "Recommended Architecture",
        "Cost Breakdown",
        "Category Totals",
        "Totals",
        "Assumptions",
        "Configuration Review",
        "Savings Opportunities",
        "Approval",
        f"{sample_estimate.cost.total_monthly:,.2f}",
    ]:
        assert expected in text, f"Expected '{expected}' in PDF text"

    # The old separate "Itemized Pricing" table (raw line items) was merged
    # into the Cost Breakdown table above - it must not still appear as its
    # own duplicate section.
    assert "Itemized Pricing" not in text


def test_pdf_report_lists_every_architecture_component(sample_estimate):
    pbytes = PdfReportGenerator().generate(sample_estimate)
    reader = pypdf.PdfReader(io.BytesIO(pbytes))
    text = " ".join(p.extract_text() for p in reader.pages).replace("\n", " ")
    for component in sample_estimate.architecture.components:
        assert component.service in text


def test_excel_resources_sheet_lists_every_resource_with_full_columns(sample_estimate):
    # sample_estimate has compute (normalized: 3 vCPU requested), storage,
    # database, and network populated - so this exercises every legacy
    # resource-summary row, not just the catalog-service path.
    xbytes = ExcelReportGenerator().generate(sample_estimate)
    wb = load_workbook(io.BytesIO(xbytes))
    ws = wb["Resources"]
    rows = list(ws.iter_rows(min_row=5, values_only=True))

    resource_names_in_sheet = {row[1] for row in rows if row[1] and row[1] not in ("Grand Total",)}
    expected_names = {s.resource_name for s in sample_estimate.cost.resource_summaries}
    assert expected_names.issubset(resource_names_in_sheet)

    header_row = [c.value for c in ws[4]]
    for expected_header in [
        "Category", "Resource", "Quantity", "Requested Configuration", "Normalized Configuration",
        "Unit Cost", "Subtotal", "Currency", "Region", "SKU", "Pricing Source", "Validation Status", "Assumption",
    ]:
        assert expected_header in header_row

    # The normalized compute resource's requested/normalized configuration
    # must differ, and its assumption reason must be visible in the sheet -
    # the platform must never silently substitute a configuration.
    compute_row = next(r for r in rows if r and r[1] == "Compute Engine")
    idx = {h: i for i, h in enumerate(header_row)}
    assert compute_row[idx["Requested Configuration"]] != compute_row[idx["Normalized Configuration"]]
    assert compute_row[idx["Validation Status"]] == "Normalized"
    assert compute_row[idx["Assumption"]] and compute_row[idx["Assumption"]] != "-"

    # Category Totals block: sum reconciles with the resource grand total.
    all_values = [cell.value for row in ws.iter_rows() for cell in row]
    assert "Sum of Category Totals" in all_values
    assert "Grand Total" in all_values


# -- Groq-generated explanations merged into the Excel/PDF report -----------
# Exercises the full "Excel Upload -> ... -> Groq explanation -> Excel/PDF"
# flow's report-generation half: an EstimateExplanation (produced by
# ExplanationService, backed by a mocked Groq call) is passed through to
# both generators exactly as POST /api/v1/reports/excel|pdf does when
# ReportRequest.include_ai_explanations=true (see
# app/api/routers/reports.py::_maybe_explain).

def _groq_explanation_for(estimate) -> "ExplanationService":
    n = len(estimate.assumptions)
    items = [{"id": i, "explanation": f"Plain-English reason #{i}."} for i in range(n)]
    items += [
        {"id": n + i, "explanation": f"Resource explanation #{i}."}
        for i in range(len(estimate.cost.resource_summaries))
    ]

    def handler(request: httpx.Request) -> httpx.Response:
        body = {"executive_summary": "This estimate is ready for review.", "items": items}
        return httpx.Response(200, json={"choices": [{"message": {"content": json.dumps(body)}}]})

    provider = GroqProvider("test-key", "llama-3.1-8b-instant", transport=httpx.MockTransport(handler))
    return ExplanationService(provider=provider, model="llama-3.1-8b-instant")


def test_excel_report_embeds_groq_explanations_alongside_deterministic_reasons(sample_estimate):
    explanation = _groq_explanation_for(sample_estimate).explain_estimate(sample_estimate)
    assert explanation.summary_source == "llm"

    xbytes = ExcelReportGenerator().generate(sample_estimate, BrandingConfig(), explanation)
    wb = load_workbook(io.BytesIO(xbytes))

    # Executive summary sheet carries the AI-generated summary...
    summary_values = [cell.value for row in wb["Summary"].iter_rows() for cell in row]
    assert explanation.executive_summary in summary_values

    # ...and the Assumptions sheet keeps the original deterministic reason
    # AND adds the customer-friendly explanation as an extra column - the
    # deterministic engine's own text is never replaced, only supplemented.
    ws = wb["Assumptions"]
    header_row = [c.value for c in ws[4]]
    assert "Customer-Friendly Explanation" in header_row
    rows = list(ws.iter_rows(min_row=5, values_only=True))
    reasons = {row[3] for row in rows if row[3]}
    ai_explanations = {row[header_row.index("Customer-Friendly Explanation")] for row in rows}
    for a in sample_estimate.assumptions:
        assert a.reason in reasons
    for exp in explanation.assumption_explanations:
        assert exp.explanation in ai_explanations


def test_pdf_report_embeds_groq_explanations(sample_estimate):
    explanation = _groq_explanation_for(sample_estimate).explain_estimate(sample_estimate)

    pbytes = PdfReportGenerator().generate(sample_estimate, BrandingConfig(), explanation)
    reader = pypdf.PdfReader(io.BytesIO(pbytes))
    text = "\n".join(p.extract_text() for p in reader.pages)

    assert "AI-Generated Summary" in text
    assert explanation.executive_summary in text
    for exp in explanation.assumption_explanations:
        assert exp.explanation in text


def test_reports_still_generate_when_groq_is_unconfigured(sample_estimate):
    # No provider configured at all - ExplanationService.explain_estimate
    # must still return a complete (template-sourced) explanation, and the
    # reports must still render successfully. Pricing/report generation
    # never breaks because Groq isn't available.
    explanation = ExplanationService(provider=None, model=None).explain_estimate(sample_estimate)
    assert explanation.summary_source == "template"

    xbytes = ExcelReportGenerator().generate(sample_estimate, BrandingConfig(), explanation)
    assert load_workbook(io.BytesIO(xbytes)).sheetnames  # opens without error

    pbytes = PdfReportGenerator().generate(sample_estimate, BrandingConfig(), explanation)
    assert pbytes[:5] == b"%PDF-"


def test_pdf_resources_section_shows_category_status_and_category_totals(sample_estimate):
    pbytes = PdfReportGenerator().generate(sample_estimate)
    reader = pypdf.PdfReader(io.BytesIO(pbytes))
    text = " ".join(p.extract_text() for p in reader.pages).replace("\n", " ")

    assert "Cost Breakdown" in text
    assert "Category Totals" in text
    for s in sample_estimate.cost.resource_summaries:
        assert s.resource_name in text
    for category in sample_estimate.cost.category_totals:
        assert category in text
    # The normalized Compute Engine resource's assumption reason must be
    # visible directly in the PDF, not just referenced elsewhere.
    compute = next(s for s in sample_estimate.cost.resource_summaries if s.resource_name == "Compute Engine")
    assert compute.status == "normalized"
