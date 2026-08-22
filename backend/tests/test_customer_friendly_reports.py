"""
Tests for the customer-friendly report content pass: the Summary/Assumptions/
Validation sections must lead with pricing and plain-English explanations,
never raw internal jargon (dotted field codes, "Validation Blockers" counts,
"Audit Log Entries" counts, raw BLOCKER/WARNING labels) - see
app/domain/assumption.py::humanize_field_code and the _build_summary /
_build_assumptions / _validation_section rewrites in the report generators.
"""
import io

import pytest
from openpyxl import load_workbook

from app.catalog.mock_provider import MockCatalogProvider
from app.core.config import Settings
from app.domain.assumption import humanize_field_code
from app.domain.branding import BrandingConfig
from app.domain.enums import MachineFamily, Region
from app.domain.requirements import ComputeRequirement, CustomerRequirement
from app.pricing.mock_provider import MockGCPPricingProvider
from app.reports.excel_generator import ExcelReportGenerator
from app.reports.pdf_generator import PdfReportGenerator
from app.services.estimation_service import EstimationService

pypdf = pytest.importorskip("pypdf")

JARGON_TERMS = ["Validation Blockers", "Validation Warnings", "Audit Log Entries", "Quality Indicator"]


@pytest.fixture
def clean_estimate():
    """An estimate with zero assumptions and zero validation findings - an
    exact e2-standard-4 match needs no substitution and passes every rule.
    Used to exercise the "everything's fine" customer-facing wording."""
    catalog = MockCatalogProvider()
    pricing = MockGCPPricingProvider()
    settings = Settings(default_normalization_strategy="balanced", default_tax_rate_percent=0, support_plan_percent=0)
    service = EstimationService(catalog=catalog, pricing_provider=pricing, settings=settings)
    req = CustomerRequirement(
        project_name="Clean Config Test",
        region=Region.US_CENTRAL1,
        compute=ComputeRequirement(machine_family=MachineFamily.E2, vcpu=4, ram_gb=16, instance_count=1),
    )
    return service.generate_estimate(req)


# -- humanize_field_code -------------------------------------------------------

def test_humanize_field_code_maps_known_codes_to_plain_labels():
    assert humanize_field_code("compute.vcpu") == "CPU count (vCPUs)"
    assert humanize_field_code("compute.ram_gb") == "Memory (RAM)"
    assert humanize_field_code("region") == "Region"
    assert humanize_field_code("database.tier") == "Database tier"


def test_humanize_field_code_falls_back_gracefully_for_unmapped_codes():
    # No entry in _FIELD_LABELS for this - must still produce something
    # readable rather than crash or leak the raw dotted/bracketed code.
    assert humanize_field_code("storage.additional_disks[0].disk_type") == "Disk type"
    assert humanize_field_code("some_new_field") == "Some new field"


# -- Excel Summary sheet: no internal jargon, pricing-first -------------------

def test_excel_summary_has_no_internal_jargon(sample_estimate):
    xbytes = ExcelReportGenerator().generate(sample_estimate, BrandingConfig())
    wb = load_workbook(io.BytesIO(xbytes))
    values = [cell.value for row in wb["Summary"].iter_rows() for cell in row]
    for term in JARGON_TERMS:
        assert term not in values, f"'{term}' should not appear on the customer-facing Summary sheet"


def test_excel_summary_explains_assumptions_in_plain_language(sample_estimate):
    assert sample_estimate.assumptions  # sanity: this estimate has assumptions to explain
    xbytes = ExcelReportGenerator().generate(sample_estimate, BrandingConfig())
    wb = load_workbook(io.BytesIO(xbytes))
    values = [cell.value for row in wb["Summary"].iter_rows() for cell in row]
    note = next((v for v in values if isinstance(v, str) and "could not be matched exactly" in v), None)
    assert note is not None
    assert str(len(sample_estimate.assumptions)) in note
    assert "Assumptions" in note  # points the reader at the detail sheet


def test_excel_summary_reassures_when_configuration_is_clean(clean_estimate):
    xbytes = ExcelReportGenerator().generate(clean_estimate, BrandingConfig())
    wb = load_workbook(io.BytesIO(xbytes))
    values = [cell.value for row in wb["Summary"].iter_rows() for cell in row]
    assert any(isinstance(v, str) and "no changes were needed" in v for v in values)


def test_excel_validation_and_audit_sheets_still_exist_as_detail_sheets(sample_estimate):
    # Per the agreed scope: hide jargon from the Summary teaser, but keep
    # the full Validation/Audit sheets intact for anyone who needs them.
    xbytes = ExcelReportGenerator().generate(sample_estimate, BrandingConfig())
    wb = load_workbook(io.BytesIO(xbytes))
    assert "Validation" in wb.sheetnames
    assert "Audit" in wb.sheetnames
    assert wb["Audit"].max_row > 4  # actual audit rows are still there, just not teased on Summary


# -- Excel/PDF Assumptions: humanized field labels -----------------------------

def test_excel_assumptions_sheet_uses_plain_language_headers_and_labels(sample_estimate):
    xbytes = ExcelReportGenerator().generate(sample_estimate, BrandingConfig())
    wb = load_workbook(io.BytesIO(xbytes))
    ws = wb["Assumptions"]
    header_row = [c.value for c in ws[4]]
    assert header_row[:4] == ["What Changed", "You Requested", "We Used", "Why"]

    field_values = {row[0] for row in ws.iter_rows(min_row=5, values_only=True) if row[0]}
    # Raw dotted codes must never leak into the customer-facing column.
    assert not any("." in str(v) for v in field_values)
    assert "CPU count (vCPUs)" in field_values


def test_pdf_assumptions_section_uses_plain_language_headers_and_labels(sample_estimate):
    pbytes = PdfReportGenerator().generate(sample_estimate)
    reader = pypdf.PdfReader(io.BytesIO(pbytes))
    text = " ".join(p.extract_text() for p in reader.pages).replace("\n", " ")
    assert "What Changed" in text
    assert "You Requested" in text
    assert "We Used" in text
    assert "CPU count (vCPUs)" in text
    assert "compute.vcpu" not in text


# -- PDF Configuration Review (formerly "Validation Results") -----------------

def test_pdf_configuration_review_softens_severity_labels(sample_estimate):
    assert sample_estimate.validation.warning_count > 0  # sanity
    pbytes = PdfReportGenerator().generate(sample_estimate)
    reader = pypdf.PdfReader(io.BytesIO(pbytes))
    text = " ".join(p.extract_text() for p in reader.pages).replace("\n", " ")
    assert "Configuration Review" in text
    assert "Worth Reviewing" in text
    assert "WARNING" not in text
    assert "BLOCKER" not in text


def test_pdf_shows_unified_badge_and_omits_empty_sections_when_clean(clean_estimate):
    # When there are zero assumptions and zero validation findings, the two
    # (formerly near-empty) "nothing to report" sections are replaced by a
    # single inline "Validation Status: Passed" badge. Savings Opportunities
    # is judged independently - clean_estimate still gets an automatic
    # Sustained Use Discount, and that's useful information worth showing,
    # not a "something's wrong" warning that the clean badge should hide.
    assert clean_estimate.cost.discounts  # sanity: this fixture does have a discount to show
    pbytes = PdfReportGenerator().generate(clean_estimate)
    reader = pypdf.PdfReader(io.BytesIO(pbytes))
    text = " ".join(p.extract_text() for p in reader.pages).replace("\n", " ")
    assert "Validation Status: Passed" in text
    assert "Configuration Review" not in text
    assert "Assumptions" not in text
    assert "Savings Opportunities" in text
    assert "Sustained Use Discount" in text


def test_pdf_executive_summary_has_no_validated_against_rules_jargon(sample_estimate):
    pbytes = PdfReportGenerator().generate(sample_estimate)
    reader = pypdf.PdfReader(io.BytesIO(pbytes))
    text = " ".join(p.extract_text() for p in reader.pages).replace("\n", " ")
    assert "validated against" not in text.lower()
    assert "could not be matched exactly" in text


def test_pdf_executive_summary_reassures_when_clean(clean_estimate):
    pbytes = PdfReportGenerator().generate(clean_estimate)
    reader = pypdf.PdfReader(io.BytesIO(pbytes))
    text = " ".join(p.extract_text() for p in reader.pages).replace("\n", " ")
    assert "no changes were needed" in text
