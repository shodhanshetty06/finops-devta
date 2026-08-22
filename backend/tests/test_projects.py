import io

from openpyxl import load_workbook

from tests.conftest import register_and_login

COMPUTE_REQUIREMENT = {
    "project_name": "Retail Inventory API",
    "region": "us-central1",
    "normalization_strategy": "performance",
    "compute": {"machine_family": "e2", "vcpu": 3, "ram_gb": 10, "instance_count": 2},
    "storage": {"disk_type": "pd-balanced", "size_gb": 100},
}

BAD_GPU_REQUIREMENT = {
    "project_name": "Bad GPU project",
    "region": "us-central1",
    "compute": {"machine_family": "e2", "vcpu": 4, "ram_gb": 16, "gpu_type": "nvidia-tesla-a100", "gpu_count": 1},
}


def test_create_and_list_project(api_client):
    _, headers = register_and_login(api_client, "owner1@example.com")
    resp = api_client.post("/api/v1/projects", json={"name": "Acme Migration"}, headers=headers)
    assert resp.status_code == 201
    project = resp.json()
    assert project["name"] == "Acme Migration"
    assert project["latest_version"] is None

    resp = api_client.get("/api/v1/projects", headers=headers)
    assert resp.status_code == 200
    names = [p["name"] for p in resp.json()]
    assert "Acme Migration" in names


def test_projects_are_isolated_between_owners(api_client):
    _, headers_a = register_and_login(api_client, "owner-a@example.com")
    _, headers_b = register_and_login(api_client, "owner-b@example.com")

    project = api_client.post("/api/v1/projects", json={"name": "A's Project"}, headers=headers_a).json()

    # B can't see it in their list.
    resp = api_client.get("/api/v1/projects", headers=headers_b)
    assert project["id"] not in [p["id"] for p in resp.json()]

    # B can't fetch it directly either.
    resp = api_client.get(f"/api/v1/projects/{project['id']}", headers=headers_b)
    assert resp.status_code == 403
    assert resp.json()["error"] == "forbidden"


def test_admin_can_access_any_project(api_client, db_session):
    _, owner_headers = register_and_login(api_client, "owner-c@example.com")
    _, admin_headers = register_and_login(api_client, "admin1@example.com", role="admin", db_session=db_session)

    project = api_client.post("/api/v1/projects", json={"name": "C's Project"}, headers=owner_headers).json()

    resp = api_client.get(f"/api/v1/projects/{project['id']}", headers=admin_headers)
    assert resp.status_code == 200

    resp = api_client.get("/api/v1/projects", headers=admin_headers)
    assert project["id"] in [p["id"] for p in resp.json()]


def test_project_endpoints_require_authentication(api_client):
    resp = api_client.get("/api/v1/projects")
    assert resp.status_code == 401
    resp = api_client.post("/api/v1/projects", json={"name": "x"})
    assert resp.status_code == 401


def test_missing_project_returns_404(api_client):
    _, headers = register_and_login(api_client, "owner-d@example.com")
    resp = api_client.get("/api/v1/projects/999999", headers=headers)
    assert resp.status_code == 404


def test_create_estimate_version_persists_and_increments(api_client):
    _, headers = register_and_login(api_client, "owner-e@example.com")
    project = api_client.post("/api/v1/projects", json={"name": "Versioned Project"}, headers=headers).json()

    resp = api_client.post(
        f"/api/v1/projects/{project['id']}/estimates",
        json={"requirement": COMPUTE_REQUIREMENT},
        headers=headers,
    )
    assert resp.status_code == 201
    v1 = resp.json()
    assert v1["version"] == 1
    assert v1["result"]["normalized_spec"]["vcpu"] == 4  # performance strategy: 3 -> 4
    assert v1["result"]["cost"]["total_monthly"] > 0

    resp = api_client.post(
        f"/api/v1/projects/{project['id']}/estimates",
        json={"requirement": COMPUTE_REQUIREMENT},
        headers=headers,
    )
    assert resp.status_code == 201
    v2 = resp.json()
    assert v2["version"] == 2
    assert v2["estimate_id"] != v1["estimate_id"]  # each run gets a fresh estimate_id

    resp = api_client.get(f"/api/v1/projects/{project['id']}/estimates", headers=headers)
    versions = resp.json()
    assert [v["version"] for v in versions] == [2, 1]  # newest first

    resp = api_client.get(f"/api/v1/projects/{project['id']}/estimates/1", headers=headers)
    assert resp.status_code == 200
    assert resp.json()["request"]["project_name"] == "Retail Inventory API"

    resp = api_client.get(f"/api/v1/projects/{project['id']}", headers=headers)
    assert resp.json()["latest_version"] == 2


def test_estimate_version_blocked_on_validation_blocker(api_client):
    _, headers = register_and_login(api_client, "owner-f@example.com")
    project = api_client.post("/api/v1/projects", json={"name": "Blocked Project"}, headers=headers).json()

    resp = api_client.post(
        f"/api/v1/projects/{project['id']}/estimates",
        json={"requirement": BAD_GPU_REQUIREMENT},
        headers=headers,
    )
    assert resp.status_code == 422
    assert resp.json()["error"] == "validation_failed"

    # Confirm nothing was persisted.
    resp = api_client.get(f"/api/v1/projects/{project['id']}/estimates", headers=headers)
    assert resp.json() == []


def test_estimate_version_report_exports(api_client):
    _, headers = register_and_login(api_client, "owner-g@example.com")
    project = api_client.post("/api/v1/projects", json={"name": "Export Project"}, headers=headers).json()
    api_client.post(
        f"/api/v1/projects/{project['id']}/estimates",
        json={"requirement": COMPUTE_REQUIREMENT},
        headers=headers,
    )

    resp = api_client.post(
        f"/api/v1/projects/{project['id']}/estimates/1/reports/excel",
        json={"prepared_for": "Acme Co."},
        headers=headers,
    )
    assert resp.status_code == 200
    assert resp.content[:2] == b"PK"

    resp = api_client.post(
        f"/api/v1/projects/{project['id']}/estimates/1/reports/pdf",
        json={},
        headers=headers,
    )
    assert resp.status_code == 200
    assert resp.content[:5] == b"%PDF-"


def test_audit_log_hidden_from_non_admins_but_visible_to_admins(api_client, db_session):
    _, owner_headers = register_and_login(api_client, "owner-j@example.com")
    _, admin_headers = register_and_login(api_client, "admin2@example.com", role="admin", db_session=db_session)
    project = api_client.post("/api/v1/projects", json={"name": "Audited Project"}, headers=owner_headers).json()
    api_client.post(
        f"/api/v1/projects/{project['id']}/estimates",
        json={"requirement": COMPUTE_REQUIREMENT},
        headers=owner_headers,
    )

    resp = api_client.get(f"/api/v1/projects/{project['id']}/estimates/1", headers=owner_headers)
    assert resp.status_code == 200
    assert resp.json()["result"]["audit_log"]["entries"] == []

    resp = api_client.get(f"/api/v1/projects/{project['id']}/estimates/1", headers=admin_headers)
    assert resp.status_code == 200
    assert len(resp.json()["result"]["audit_log"]["entries"]) > 0


def test_excel_export_audit_sheet_hidden_from_non_admins_but_visible_to_admins(api_client, db_session):
    _, owner_headers = register_and_login(api_client, "owner-k@example.com")
    _, admin_headers = register_and_login(api_client, "admin3@example.com", role="admin", db_session=db_session)
    project = api_client.post("/api/v1/projects", json={"name": "Audited Export Project"}, headers=owner_headers).json()
    api_client.post(
        f"/api/v1/projects/{project['id']}/estimates",
        json={"requirement": COMPUTE_REQUIREMENT},
        headers=owner_headers,
    )

    resp = api_client.post(
        f"/api/v1/projects/{project['id']}/estimates/1/reports/excel",
        json={"prepared_for": "Acme Co."},
        headers=owner_headers,
    )
    assert resp.status_code == 200
    ws = load_workbook(io.BytesIO(resp.content))["Audit"]
    assert ws.max_row == 4  # title, subtitle, blank, header row - no entries
    assert ws["A2"].value.startswith("0 recorded decisions")

    resp = api_client.post(
        f"/api/v1/projects/{project['id']}/estimates/1/reports/excel",
        json={"prepared_for": "Acme Co."},
        headers=admin_headers,
    )
    assert resp.status_code == 200
    ws = load_workbook(io.BytesIO(resp.content))["Audit"]
    assert ws.max_row > 4


def test_other_user_cannot_access_projects_estimates(api_client):
    _, headers_a = register_and_login(api_client, "owner-h@example.com")
    _, headers_b = register_and_login(api_client, "owner-i@example.com")
    project = api_client.post("/api/v1/projects", json={"name": "H's Project"}, headers=headers_a).json()
    api_client.post(
        f"/api/v1/projects/{project['id']}/estimates",
        json={"requirement": COMPUTE_REQUIREMENT},
        headers=headers_a,
    )

    resp = api_client.get(f"/api/v1/projects/{project['id']}/estimates", headers=headers_b)
    assert resp.status_code == 403
