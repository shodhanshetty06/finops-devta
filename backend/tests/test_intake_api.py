"""API-level tests for Phase 4 intake endpoints - both the stateless
/api/v1/intake/* endpoints and the project-scoped persistence shortcuts."""
import io

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.api.dependencies import get_estimation_service
from app.core.exceptions import ValidationFailedError
from app.domain.enums import Severity
from app.domain.validation import ValidationResult
from app.intake.excel_template import ExcelTemplateGenerator
from app.main import app
from tests.conftest import register_and_login

client = TestClient(app)


def _filled_questionnaire_bytes(values: dict[str, object]) -> bytes:
    blank = ExcelTemplateGenerator().generate()
    wb = load_workbook(io.BytesIO(blank))
    ws = wb["Questionnaire"]
    for row in ws.iter_rows():
        label = row[0].value
        if label in values:
            row[1].value = values[label]
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


GOOD_QUESTIONNAIRE = {
    "Project Name": "API Excel Test",
    "Region": "us-central1",
    "Machine Family": "e2",
    "vCPU": 4,
    "RAM (GB)": 16,
    "Instance Count": 1,
}


def test_download_template_endpoint():
    resp = client.get("/api/v1/intake/excel/template")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert resp.content[:2] == b"PK"


def test_upload_excel_returns_parsed_requirement():
    file_bytes = _filled_questionnaire_bytes(GOOD_QUESTIONNAIRE)
    resp = client.post(
        "/api/v1/intake/excel",
        files={"file": ("questionnaire.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["requirement"]["project_name"] == "API Excel Test"
    assert body["issues"] == []
    assert body["estimate"] is None


def test_upload_excel_with_auto_estimate_returns_priced_result():
    file_bytes = _filled_questionnaire_bytes(GOOD_QUESTIONNAIRE)
    resp = client.post(
        "/api/v1/intake/excel?auto_estimate=true",
        files={"file": ("questionnaire.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["estimate"] is not None
    assert body["estimate"]["cost"]["total_monthly"] > 0


class _BlockingEstimationService:
    """Stub that always raises ValidationFailedError, so tests can exercise
    the auto-estimate-hits-a-blocker path without needing to construct a
    real requirement that fails validation end-to-end."""

    def generate_estimate(self, requirement, *, force=False, commitment_term_years=0):
        raise ValidationFailedError(
            "Request has 1 blocking validation issue(s).",
            results=[
                ValidationResult(
                    field="region", rule="region_validation", requested_value="mars-north1",
                    supported_value="us-central1", is_valid=False, severity=Severity.BLOCKER,
                    reason="'mars-north1' is not a supported region in the catalog.",
                    recommendation="Choose a supported region.",
                ),
                ValidationResult(
                    field="compute.vcpu", rule="cpu_validation", requested_value="7",
                    supported_value=None, is_valid=False, severity=Severity.WARNING,
                    reason="7 vCPU is not a standard configuration.", recommendation="Normalize it.",
                ),
            ],
        )


def test_upload_excel_auto_estimate_blocker_still_returns_parsed_data():
    """Regression test: a blocked auto-estimate must degrade gracefully -
    returning the parsed requirement plus the blocking issue(s) - instead of
    raising past the router and losing everything the parser recovered."""
    file_bytes = _filled_questionnaire_bytes(GOOD_QUESTIONNAIRE)
    app.dependency_overrides[get_estimation_service] = lambda: _BlockingEstimationService()
    try:
        resp = client.post(
            "/api/v1/intake/excel?auto_estimate=true",
            files={"file": ("questionnaire.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    finally:
        app.dependency_overrides.pop(get_estimation_service, None)

    assert resp.status_code == 200
    body = resp.json()
    assert body["requirement"]["project_name"] == "API Excel Test"
    assert body["estimate"] is None
    blocker_issues = [i for i in body["issues"] if i["severity"] == "blocker"]
    assert len(blocker_issues) == 1
    assert blocker_issues[0]["field"] == "region"
    # Non-blocker validation results (warnings) must not leak into `issues`.
    assert all(i["field"] != "compute.vcpu" for i in body["issues"])


def test_upload_unreadable_file_returns_400():
    resp = client.post(
        "/api/v1/intake/excel",
        files={"file": ("bad.xlsx", b"not a real workbook", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 400
    assert resp.json()["error"] == "intake_parse_error"


def test_text_extraction_endpoint():
    resp = client.post("/api/v1/intake/text", json={
        "project_name": "Text Test",
        "text": "500 users, HA required, 99.99% uptime, 100GB database.",
    })
    assert resp.status_code == 200
    body = resp.json()
    assert body["requirement"]["business"]["total_users"] == 500
    assert body["requirement"]["availability"]["high_availability"] is True
    assert len(body["notes"]) > 0


def test_text_extraction_with_auto_estimate():
    resp = client.post("/api/v1/intake/text?auto_estimate=true", json={
        "project_name": "Text Test Priced",
        "text": "500 users, HA required, 99.99% uptime, 100GB database.",
    })
    assert resp.status_code == 200
    body = resp.json()
    assert body["estimate"] is not None
    assert body["estimate"]["cost"]["total_monthly"] > 0
    # This request had no compute section, so the AI recommendation engine
    # should have filled it in - confirm that shows up as an assumption.
    assert any(a["strategy_applied"] == "ai_recommendation" for a in body["estimate"]["assumptions"])


def test_project_intake_excel_creates_version(api_client):
    _, headers = register_and_login(api_client, "intake-excel@example.com")
    project = api_client.post("/api/v1/projects", json={"name": "Intake Excel Project"}, headers=headers).json()

    file_bytes = _filled_questionnaire_bytes(GOOD_QUESTIONNAIRE)
    resp = api_client.post(
        f"/api/v1/projects/{project['id']}/intake/excel",
        files={"file": ("q.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )
    assert resp.status_code == 201
    body = resp.json()
    assert body["version"] == 1
    assert body["result"]["cost"]["total_monthly"] > 0


def test_project_intake_excel_unparseable_returns_400(api_client):
    _, headers = register_and_login(api_client, "intake-bad@example.com")
    project = api_client.post("/api/v1/projects", json={"name": "Bad Intake Project"}, headers=headers).json()

    # A blank Project Name no longer makes a workbook unparseable (it just
    # gets an auto-generated default - see excel_parser.py) - a genuinely
    # corrupt/unreadable file is the remaining way to hit this 400.
    resp = api_client.post(
        f"/api/v1/projects/{project['id']}/intake/excel",
        files={"file": ("not-a-workbook.xlsx", b"this is not a valid xlsx file at all",
                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )
    assert resp.status_code == 400
    assert resp.json()["error"] == "intake_parse_error"

    # Confirm nothing was persisted.
    resp = api_client.get(f"/api/v1/projects/{project['id']}/estimates", headers=headers)
    assert resp.json() == []


def test_project_intake_text_creates_version(api_client):
    _, headers = register_and_login(api_client, "intake-text@example.com")
    project = api_client.post("/api/v1/projects", json={"name": "Intake Text Project"}, headers=headers).json()

    resp = api_client.post(
        f"/api/v1/projects/{project['id']}/intake/text",
        json={"project_name": "From Text", "text": "500 users, HA required, 99.99% uptime, 100GB database."},
        headers=headers,
    )
    assert resp.status_code == 201
    body = resp.json()
    assert body["version"] == 1
    assert body["request"]["project_name"] == "From Text"
    assert body["result"]["cost"]["total_monthly"] > 0


def test_project_intake_requires_authentication(api_client):
    resp = api_client.post(
        "/api/v1/projects/1/intake/text",
        json={"project_name": "x", "text": "500 users"},
    )
    assert resp.status_code == 401
