"""
API-level tests for POST /api/v1/explanations/* (app/api/routers/explanation.py)
and the opt-in include_ai_explanations flag on the report export endpoints
(app/api/routers/reports.py). Uses TestClient + a dependency override for
get_explanation_service, same pattern as tests/test_currency.py's
get_currency_converter override, so no real Groq call is ever made."""
import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.domain.branding import BrandingConfig
from app.main import app
from app.services.explanation_service import ExplanationService, get_explanation_service

pypdf = pytest.importorskip("pypdf")


@pytest.fixture
def api_client():
    return TestClient(app)


@pytest.fixture(autouse=True)
def _no_real_groq_calls():
    # Regardless of the developer's local .env, tests never hit the real
    # Groq API - the default (no override) ExplanationService has
    # provider=None whenever GROQ_API_KEY/GROQ_MODEL aren't set in the test
    # environment, which is exactly the "missing API key" fallback path.
    yield


def test_explain_estimate_endpoint_returns_template_explanations_by_default(api_client, sample_estimate):
    resp = api_client.post("/api/v1/explanations/estimate", json={"estimate": sample_estimate.model_dump(mode="json")})
    assert resp.status_code == 200
    body = resp.json()

    assert body["summary_source"] == "template"
    assert len(body["assumption_explanations"]) == len(sample_estimate.assumptions)
    assert len(body["resource_explanations"]) == len(sample_estimate.cost.resource_summaries)
    for exp in body["assumption_explanations"]:
        assert exp["explanation"]
        assert exp["source"] == "template"


def test_explain_estimate_endpoint_with_llm_configured(api_client, sample_estimate):
    class StubService(ExplanationService):
        def __init__(self):
            super().__init__(provider=None, model="stub-model")

        def explain_estimate(self, estimate):
            result = super().explain_estimate(estimate)
            return result.model_copy(update={"summary_source": "llm"})

    app.dependency_overrides[get_explanation_service] = lambda: StubService()
    try:
        resp = api_client.post("/api/v1/explanations/estimate", json={"estimate": sample_estimate.model_dump(mode="json")})
        assert resp.status_code == 200
        assert resp.json()["summary_source"] == "llm"
    finally:
        app.dependency_overrides.clear()


def test_explain_scenario_comparison_endpoint(api_client):
    comparison_json = {
        "base": {"name": "Base", "total_monthly": 1000.0, "total_yearly": 12000.0, "currency": "USD",
                  "delta_vs_base_monthly": 0.0, "delta_vs_base_percent": 0.0},
        "scenarios": [
            {"name": "Upsize", "total_monthly": 1500.0, "total_yearly": 18000.0, "currency": "USD",
             "delta_vs_base_monthly": 500.0, "delta_vs_base_percent": 50.0},
        ],
    }
    resp = api_client.post("/api/v1/explanations/scenario-comparison", json={"comparison": comparison_json})
    assert resp.status_code == 200
    body = resp.json()
    assert body["source"] == "template"
    assert "500.00" in body["scenario_explanations"][0]["explanation"]


# -- Opt-in AI explanations on report exports ----------------------------------

def test_pdf_report_without_ai_explanations_is_unchanged_by_default(api_client, sample_estimate):
    resp = api_client.post("/api/v1/reports/pdf", json={
        "estimate": sample_estimate.model_dump(mode="json"),
        "branding": BrandingConfig().model_dump(mode="json"),
    })
    assert resp.status_code == 200
    reader = pypdf.PdfReader(io.BytesIO(resp.content))
    text = "\n".join(p.extract_text() for p in reader.pages)
    assert "AI-Generated Summary" not in text


def test_pdf_report_with_ai_explanations_adds_ai_content(api_client, sample_estimate):
    resp = api_client.post("/api/v1/reports/pdf", json={
        "estimate": sample_estimate.model_dump(mode="json"),
        "branding": BrandingConfig().model_dump(mode="json"),
        "include_ai_explanations": True,
    })
    assert resp.status_code == 200
    reader = pypdf.PdfReader(io.BytesIO(resp.content))
    text = "\n".join(p.extract_text() for p in reader.pages)
    assert "AI-Generated Summary" in text
    assert "Customer-Friendly Explanation" in text


def test_excel_report_with_ai_explanations_adds_extra_column(api_client, sample_estimate):
    resp = api_client.post("/api/v1/reports/excel", json={
        "estimate": sample_estimate.model_dump(mode="json"),
        "branding": BrandingConfig().model_dump(mode="json"),
        "include_ai_explanations": True,
    })
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb["Assumptions"]
    header_row = [c.value for c in ws[4]]
    assert "Customer-Friendly Explanation" in header_row


def test_excel_report_without_ai_explanations_has_no_extra_column(api_client, sample_estimate):
    resp = api_client.post("/api/v1/reports/excel", json={
        "estimate": sample_estimate.model_dump(mode="json"),
        "branding": BrandingConfig().model_dump(mode="json"),
    })
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb["Assumptions"]
    header_row = [c.value for c in ws[4]]
    assert "Customer-Friendly Explanation" not in header_row
