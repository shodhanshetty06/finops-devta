"""End-to-end test for the full "Excel Upload -> Parse -> Validate/Normalize
-> Pricing Engine -> Groq explanation -> Excel/PDF report" pipeline
described in the platform's intake/report design:

    Excel Upload -> Parse Excel -> Validate/Normalize -> Pricing Engine
    -> structured results -> Groq customer-friendly explanations
    -> final Excel + PDF report

Every stage reuses the platform's real components (ExcelQuestionnaireParser,
EstimationService -> ValidationRuleEngine/NormalizationEngine/PricingEngine,
ExplanationService, ExcelReportGenerator/PdfReportGenerator) - nothing here
mocks the deterministic pipeline itself, only the outbound Groq HTTP call
(httpx.MockTransport, same pattern as tests/test_llm_groq_provider.py). This
confirms Groq only ever adds prose on top of numbers the deterministic
engines already computed, and that currency/target-currency handling and
the Groq-unavailable fallback both work through the complete pipeline, not
just in isolation.
"""
import io
import json

import httpx
from openpyxl import load_workbook

from app.domain.branding import BrandingConfig
from app.intake.excel_parser import ExcelQuestionnaireParser
from app.intake.excel_template import ExcelTemplateGenerator
from app.llm.groq_provider import GroqProvider
from app.pricing.currency_converter import CurrencyConverter, CurrencyExchangeClient, InMemoryCurrencyRateCache, convert_estimate_currency
from app.reports.excel_generator import ExcelReportGenerator
from app.reports.pdf_generator import PdfReportGenerator
from app.services.explanation_service import ExplanationService
from app.services.estimation_service import EstimationService


def _filled_questionnaire_bytes() -> bytes:
    """A questionnaire requesting 3 vCPU - not a directly supported
    machine-type tier - so the pipeline is forced to normalize it, exactly
    like the "requested 3 vCPU, assumed 2 vCPU" example in the platform spec."""
    wb = load_workbook(io.BytesIO(ExcelTemplateGenerator().generate()))
    ws = wb["Questionnaire"]
    values = {
        "Project Name": "Retail Inventory API",
        "Region": "us-central1",
        "Machine Family": "e2",
        "vCPU": 3,
        "RAM (GB)": 10,
        "Instance Count": 1,
        "Disk Type": "pd-balanced",
        "Size (GB)": 100,
        "Snapshot Enabled": "FALSE",
    }
    for row in ws.iter_rows():
        label = row[0].value
        if label in values:
            row[1].value = values[label]
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _groq_handler(request: httpx.Request) -> httpx.Response:
    sent = json.loads(json.loads(request.content)["messages"][1]["content"])
    items = [
        {"id": item["id"], "explanation": (
            f'The requested {item["field"]} value of "{item["requested_value"]}" is not available in the '
            f'supported pricing catalog. For this estimate, we used "{item["used_value"]}" instead.'
        )}
        for item in sent["assumptions"]
    ]
    items += [{"id": r["id"], "explanation": f'{r["resource_name"]} is priced at {r["currency"]} {r["subtotal_monthly"]:,.2f}/mo.'}
              for r in sent["resources"]]
    body = {"executive_summary": "Your estimate is ready - see the breakdown below.", "items": items}
    return httpx.Response(200, json={"choices": [{"message": {"content": json.dumps(body)}}]})


def test_full_pipeline_excel_upload_through_groq_explained_reports(catalog, pricing_provider, settings):
    # 1) Excel Upload + Parse
    requirement, issues = ExcelQuestionnaireParser().parse(_filled_questionnaire_bytes())
    assert issues == []
    assert requirement.compute.vcpu == 3

    # 2) Validate / Normalize / Price - the deterministic pricing engine
    #    remains the sole source of truth for every number produced here.
    estimate = EstimationService(catalog=catalog, pricing_provider=pricing_provider, settings=settings).generate_estimate(requirement)
    vcpu_assumption = next(a for a in estimate.assumptions if a.field == "compute.vcpu")
    assert vcpu_assumption.requested_value == "3"
    assert vcpu_assumption.used_value != "3"  # substituted to a supported tier, never silently dropped

    # 3) Groq customer-friendly explanations (mocked - no real network call)
    provider = GroqProvider("test-key", "llama-3.1-8b-instant", transport=httpx.MockTransport(_groq_handler))
    explanation = ExplanationService(provider=provider, model="llama-3.1-8b-instant").explain_estimate(estimate)
    assert explanation.summary_source == "llm"
    vcpu_explanation = next(
        e for e, a in zip(explanation.assumption_explanations, estimate.assumptions) if a.field == "compute.vcpu"
    )
    # Groq only phrased the substitution - it must reference the exact
    # requested/used values the engine decided, and no technical jargon.
    assert "3" in vcpu_explanation.explanation
    assert vcpu_assumption.used_value in vcpu_explanation.explanation
    for banned_term in ("normalization", "fallback", "strategy_applied"):
        assert banned_term not in vcpu_explanation.explanation.lower()

    # 4) Final Excel + PDF report, in the estimate's native currency
    xbytes = ExcelReportGenerator().generate(estimate, BrandingConfig(), explanation)
    wb = load_workbook(io.BytesIO(xbytes))
    summary_values = [cell.value for row in wb["Summary"].iter_rows() for cell in row]
    assert explanation.executive_summary in summary_values
    assert estimate.cost.total_monthly in summary_values
    assert estimate.cost.currency in [v for row in wb["Summary"].iter_rows() for v in (c.value for c in row)]

    pbytes = PdfReportGenerator().generate(estimate, BrandingConfig(), explanation)
    assert pbytes[:5] == b"%PDF-"


def test_full_pipeline_falls_back_to_deterministic_text_when_groq_unavailable(catalog, pricing_provider, settings):
    # Same pipeline, but Groq is entirely unconfigured (provider=None, the
    # exact state of a deployment with no GROQ_API_KEY set). Pricing and
    # report generation must still complete successfully end to end.
    requirement, issues = ExcelQuestionnaireParser().parse(_filled_questionnaire_bytes())
    assert issues == []

    estimate = EstimationService(catalog=catalog, pricing_provider=pricing_provider, settings=settings).generate_estimate(requirement)
    explanation = ExplanationService(provider=None, model=None).explain_estimate(estimate)
    assert explanation.summary_source == "template"
    assert all(a.source == "template" for a in explanation.assumption_explanations)

    xbytes = ExcelReportGenerator().generate(estimate, BrandingConfig(), explanation)
    assert load_workbook(io.BytesIO(xbytes)).sheetnames

    pbytes = PdfReportGenerator().generate(estimate, BrandingConfig(), explanation)
    assert pbytes[:5] == b"%PDF-"


def test_full_pipeline_renders_report_in_user_selected_currency(catalog, pricing_provider, settings):
    # The report must render in whatever currency the user selected
    # (target_currency) - converted for display only, via the real
    # CurrencyConverter (mocked HTTP), never re-priced or invented by Groq.
    requirement, issues = ExcelQuestionnaireParser().parse(_filled_questionnaire_bytes())
    assert issues == []
    estimate = EstimationService(catalog=catalog, pricing_provider=pricing_provider, settings=settings).generate_estimate(requirement)
    assert estimate.cost.currency == "USD"

    def rate_handler(request: httpx.Request) -> httpx.Response:
        return httpx.Response(200, json={"amount": 1.0, "base": "USD", "date": "2026-08-08", "rates": {"INR": 83.0}})

    converter = CurrencyConverter(
        CurrencyExchangeClient("https://api.frankfurter.dev/v1/latest", transport=httpx.MockTransport(rate_handler)),
        InMemoryCurrencyRateCache(), ttl_seconds=60, symbols=["INR"],
    )
    converted = convert_estimate_currency(estimate, "INR", converter)
    assert converted.cost.currency == "INR"
    assert converted.cost.total_monthly == round(estimate.cost.total_monthly * 83.0, 2)

    provider = GroqProvider("test-key", "llama-3.1-8b-instant", transport=httpx.MockTransport(_groq_handler))
    explanation = ExplanationService(provider=provider, model="llama-3.1-8b-instant").explain_estimate(converted)
    # Groq was given the already-converted INR figures, never asked to convert anything itself.
    for r in explanation.resource_explanations:
        assert "USD" not in r.explanation

    xbytes = ExcelReportGenerator().generate(converted, BrandingConfig(), explanation)
    wb = load_workbook(io.BytesIO(xbytes))
    all_values = [cell.value for row in wb["Summary"].iter_rows() for cell in row]
    assert "INR" in all_values
